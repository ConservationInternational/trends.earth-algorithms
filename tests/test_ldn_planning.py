"""Unit tests for te_algorithms.gdal.land_deg.ldn_planning.

Uses small synthetic rasters created in memory so no GDAL file I/O is needed
for the algorithmic logic (BAU), and temp files for the raster-writing tests.
"""

import json
import os
import tempfile
import unittest

import numpy as np
import pytest

pytest.importorskip("osgeo.gdal", reason="GDAL not available")
pytest.importorskip("osgeo.ogr", reason="OGR not available")
pytest.importorskip("osgeo.osr", reason="OSR not available")

from osgeo import gdal, osr  # noqa: E402

from te_algorithms.gdal.land_deg import ldn_planning  # noqa: E402


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _write_raster(arr: np.ndarray, path: str, nodata: int = -32768) -> None:
    """Write a single-band Int16 raster to *path*."""
    gdal.UseExceptions()
    driver = gdal.GetDriverByName("GTiff")
    ysize, xsize = arr.shape
    ds = driver.Create(path, xsize, ysize, 1, gdal.GDT_Int16)
    # Small WGS84 extent centred on 0°N 0°E
    gt = (0.0, 0.1, 0.0, 5.0, 0.0, -0.1)  # 10 cols × 10 rows → 1°×1°
    ds.SetGeoTransform(gt)
    srs = osr.SpatialReference()
    srs.ImportFromEPSG(4326)
    ds.SetProjection(srs.ExportToWkt())
    band = ds.GetRasterBand(1)
    band.SetNoDataValue(nodata)
    band.WriteArray(arr)
    ds.FlushCache()
    del ds


# ---------------------------------------------------------------------------
# ARR classification tests
# ---------------------------------------------------------------------------


class TestARRClassification(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def _status_path(self, arr: np.ndarray) -> str:
        path = os.path.join(self.tmpdir, "status.tif")
        _write_raster(arr, path)
        return path

    def test_3class_all_degraded(self):
        arr = np.full((5, 5), -1, dtype=np.int16)
        out = os.path.join(self.tmpdir, "arr_deg.tif")
        summary, path = ldn_planning.compute_arr_classification(
            self._status_path(arr), output_path=out
        )
        self.assertEqual(path, out)
        self.assertGreater(summary["reverse_km2"], 0)
        self.assertAlmostEqual(summary["avoid_km2"], 0.0, places=3)
        self.assertAlmostEqual(summary["reduce_km2"], 0.0, places=3)

    def test_3class_all_stable(self):
        arr = np.full((5, 5), 0, dtype=np.int16)
        out = os.path.join(self.tmpdir, "arr_stab.tif")
        summary, path = ldn_planning.compute_arr_classification(
            self._status_path(arr), output_path=out
        )
        # No trajectory → all non-degraded = Avoid
        self.assertGreater(summary["avoid_km2"], 0)
        self.assertAlmostEqual(summary["reduce_km2"], 0.0, places=3)
        self.assertAlmostEqual(summary["reverse_km2"], 0.0, places=3)

    def test_3class_mixed(self):
        arr = np.array(
            [
                [-1, 0, 1, 0, -1],
                [0, -1, 0, 1, 0],
                [1, 0, -1, 0, 1],
                [0, 1, 0, -1, 0],
                [-1, 0, 1, 0, -1],
            ],
            dtype=np.int16,
        )
        out = os.path.join(self.tmpdir, "arr_mix.tif")
        summary, path = ldn_planning.compute_arr_classification(
            self._status_path(arr), output_path=out
        )
        self.assertGreater(summary["reverse_km2"], 0)
        self.assertGreater(summary["avoid_km2"], 0)
        self.assertAlmostEqual(
            summary["avoid_km2"] + summary["reverse_km2"],
            summary["total_km2"],
            places=2,
        )

    def test_7class_degraded_values(self):
        # 7-class: 1=persistent deg, 2=recent deg, 3=baseline deg → all Reverse
        arr = np.array(
            [
                [1, 2, 3, 4, 5],
                [6, 7, 4, 1, 2],
                [3, 4, 5, 6, 7],
                [1, 3, 5, 7, 4],
                [2, 4, 6, 1, 3],
            ],
            dtype=np.int16,
        )
        out = os.path.join(self.tmpdir, "arr_7cls.tif")
        summary, path = ldn_planning.compute_arr_classification(
            self._status_path(arr), output_path=out
        )
        self.assertGreater(summary["reverse_km2"], 0)
        self.assertEqual(summary["status_type"], "7class")

    def test_with_trajectory_adds_reduce(self):
        status = np.full((5, 5), 0, dtype=np.int16)  # all stable → Avoid without LPD
        # LPD class 3 = "stable but stressed": not degraded, but at risk → Reduce.
        # (Classes 1/2 are already degraded and would be Reverse, so they cannot
        # supply the at-risk signal.)
        traj = np.full((5, 5), 3, dtype=np.int16)

        s_path = os.path.join(self.tmpdir, "status_t.tif")
        t_path = os.path.join(self.tmpdir, "traj.tif")
        out = os.path.join(self.tmpdir, "arr_traj.tif")
        _write_raster(status, s_path)
        _write_raster(traj, t_path)
        summary, _ = ldn_planning.compute_arr_classification(
            s_path, output_path=out, trajectory_path=t_path
        )
        # With stressed LPD, all non-degraded stable → Reduce
        self.assertGreater(summary["reduce_km2"], 0)
        self.assertAlmostEqual(summary["avoid_km2"], 0.0, places=3)

    def test_lpd_declining_is_not_at_risk_signal(self):
        # LPD class 1 (Declining) is already degraded; when overall status is
        # non-degraded, class 1 must NOT create Reduce pixels (it is not a valid
        # non-degraded early-warning class). All should remain Avoid.
        status = np.full((5, 5), 0, dtype=np.int16)  # not degraded overall
        traj = np.full((5, 5), 1, dtype=np.int16)  # Declining (already-degraded class)

        s_path = os.path.join(self.tmpdir, "status_d.tif")
        t_path = os.path.join(self.tmpdir, "traj_d.tif")
        out = os.path.join(self.tmpdir, "arr_traj_d.tif")
        _write_raster(status, s_path)
        _write_raster(traj, t_path)
        summary, _ = ldn_planning.compute_arr_classification(
            s_path, output_path=out, trajectory_path=t_path
        )
        self.assertAlmostEqual(summary["reduce_km2"], 0.0, places=3)
        self.assertGreater(summary["avoid_km2"], 0)

    def test_nodata_excluded_from_summary(self):
        arr = np.full((5, 5), -32768, dtype=np.int16)
        out = os.path.join(self.tmpdir, "arr_nd.tif")
        summary, _ = ldn_planning.compute_arr_classification(
            self._status_path(arr), output_path=out
        )
        self.assertAlmostEqual(summary["total_km2"], 0.0, places=3)
        self.assertGreater(summary["nodata_km2"], 0)


# ---------------------------------------------------------------------------
# BAU rate tests
# ---------------------------------------------------------------------------


class TestBAURate(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def test_no_change(self):
        arr = np.array(
            [[-1, 0, 0], [0, -1, 0], [0, 0, -1]],
            dtype=np.int16,
        )
        p = os.path.join(self.tmpdir, "status_bau.tif")
        _write_raster(arr, p)
        summary = ldn_planning.compute_bau_rate(
            p, 1, p, 1, year_initial=2000, year_final=2015, target_year=2030
        )
        self.assertAlmostEqual(summary["annual_change_km2"], 0.0, places=3)
        # BAU projection = reporting degraded (no change)
        proj_key = "bau_projection_2030_km2"
        self.assertAlmostEqual(
            summary[proj_key], summary["degraded_area_reporting_km2"], places=2
        )
        # Target = baseline
        self.assertAlmostEqual(
            summary["ldntarget_km2"], summary["degraded_area_baseline_km2"], places=2
        )

    def test_growing_degradation(self):
        # Baseline: 3 degraded pixels; reporting: 6 degraded pixels
        baseline = np.array(
            [[-1, 0, 0], [0, -1, 0], [0, 0, -1]],
            dtype=np.int16,
        )
        reporting = np.array(
            [[-1, -1, 0], [-1, -1, 0], [0, -1, -1]],
            dtype=np.int16,
        )
        p_bl = os.path.join(self.tmpdir, "bau_bl.tif")
        p_rp = os.path.join(self.tmpdir, "bau_rp.tif")
        _write_raster(baseline, p_bl)
        _write_raster(reporting, p_rp)
        summary = ldn_planning.compute_bau_rate(
            p_bl, 1, p_rp, 1, year_initial=2000, year_final=2015, target_year=2030
        )
        # Should have positive annual change
        self.assertGreater(summary["annual_change_km2"], 0)
        # Shortfall should be > 0
        self.assertGreater(summary["shortfall_km2"], 0)
        # Target = baseline degraded area
        self.assertAlmostEqual(
            summary["ldntarget_km2"], summary["degraded_area_baseline_km2"], places=2
        )

    def test_no_reporting_layer(self):
        arr = np.array([[-1, 0], [0, -1]], dtype=np.int16)
        p = os.path.join(self.tmpdir, "bau_norepobj.tif")
        _write_raster(arr, p)
        summary = ldn_planning.compute_bau_rate(p, 1)
        self.assertIsNone(summary["annual_change_km2"])
        self.assertIsNone(summary["degraded_area_reporting_km2"])


# ---------------------------------------------------------------------------
# ARR scenario tests
# ---------------------------------------------------------------------------


class TestApplyScenario(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        # 5×5 ARR raster: rows 0-1 = Reverse, 2-3 = Reduce, 4 = Avoid
        arr = np.array(
            [
                [3, 3, 3, 3, 3],
                [3, 3, 3, 3, 3],
                [2, 2, 2, 2, 2],
                [2, 2, 2, 2, 2],
                [1, 1, 1, 1, 1],
            ],
            dtype=np.int16,
        )
        self.arr_path = os.path.join(self.tmpdir, "arr.tif")
        _write_raster(arr, self.arr_path, nodata=-32768)

    def _full_wkt(self):
        # WKT covering the full 5×5 raster extent (0°N 0°E to 0.5°N 0.5°E)
        return "POLYGON ((0 0, 0.5 0, 0.5 5.0, 0 5.0, 0 0))"

    def test_reverse_generates_gains_only(self):
        targets = [
            {
                "wkt_geometry": self._full_wkt(),
                "intervention": "reverse",
                "effectiveness": 1.0,
            }
        ]
        out = os.path.join(self.tmpdir, "scen_rev.tif")
        summary, path = ldn_planning.apply_scenario(
            self.arr_path, targets=targets, output_path=out
        )
        self.assertGreater(summary["gains_km2_reverse"], 0)
        self.assertAlmostEqual(summary["avoided_losses_km2_reduce"], 0.0, places=3)
        self.assertAlmostEqual(summary["avoided_losses_km2_avoid"], 0.0, places=3)

    def test_reduce_generates_avoided_losses_not_gains(self):
        targets = [
            {
                "wkt_geometry": self._full_wkt(),
                "intervention": "reduce",
                "effectiveness": 1.0,
            }
        ]
        out = os.path.join(self.tmpdir, "scen_red.tif")
        summary, _ = ldn_planning.apply_scenario(
            self.arr_path, targets=targets, output_path=out
        )
        self.assertAlmostEqual(summary["gains_km2_reverse"], 0.0, places=3)
        self.assertGreater(summary["avoided_losses_km2_reduce"], 0)
        self.assertAlmostEqual(summary["total_gains_km2"], 0.0, places=3)

    def test_avoid_generates_avoided_losses_not_gains(self):
        targets = [
            {
                "wkt_geometry": self._full_wkt(),
                "intervention": "avoid",
                "effectiveness": 1.0,
            }
        ]
        out = os.path.join(self.tmpdir, "scen_av.tif")
        summary, _ = ldn_planning.apply_scenario(
            self.arr_path, targets=targets, output_path=out
        )
        self.assertAlmostEqual(summary["gains_km2_reverse"], 0.0, places=3)
        self.assertGreater(summary["avoided_losses_km2_avoid"], 0)

    def test_effectiveness_zero_yields_no_effect(self):
        targets = [
            {
                "wkt_geometry": self._full_wkt(),
                "intervention": "reverse",
                "effectiveness": 0.0,
            }
        ]
        out = os.path.join(self.tmpdir, "scen_eff0.tif")
        summary, _ = ldn_planning.apply_scenario(
            self.arr_path, targets=targets, output_path=out
        )
        self.assertAlmostEqual(summary["total_gains_km2"], 0.0, places=3)

    def test_empty_targets(self):
        out = os.path.join(self.tmpdir, "scen_empty.tif")
        summary, _ = ldn_planning.apply_scenario(
            self.arr_path, targets=[], output_path=out
        )
        self.assertAlmostEqual(summary["total_gains_km2"], 0.0, places=3)
        self.assertAlmostEqual(summary["total_avoided_losses_km2"], 0.0, places=3)


# ---------------------------------------------------------------------------
# Report writer tests
# ---------------------------------------------------------------------------


class TestReportWriter(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def test_excel_written(self):
        from te_algorithms.gdal.land_deg.ldn_planning_report import (
            save_ldn_planning_excel,
        )
        from pathlib import Path

        arr_summary = {
            "avoid_km2": 10.0,
            "reduce_km2": 5.0,
            "reverse_km2": 3.0,
            "total_km2": 18.0,
        }
        bau_summary = {
            "year_initial": 2000,
            "year_final": 2015,
            "target_year": 2030,
            "total_area_km2": 18.0,
            "degraded_area_baseline_km2": 3.0,
            "pct_degraded_baseline": 16.7,
            "degraded_area_reporting_km2": 4.0,
            "pct_degraded_reporting": 22.2,
            "annual_change_km2": 0.067,
            "bau_projection_2030_km2": 5.0,
            "ldntarget_km2": 3.0,
            "shortfall_km2": 2.0,
        }
        out = Path(self.tmpdir) / "test_report.xlsx"
        save_ldn_planning_excel(out, arr_summary=arr_summary, bau_summary=bau_summary)
        self.assertTrue(out.exists())
        self.assertGreater(out.stat().st_size, 0)

    def test_json_written(self):
        from te_algorithms.gdal.land_deg.ldn_planning_report import (
            save_ldn_planning_json,
        )
        from pathlib import Path

        arr = {"avoid_km2": 10.0, "reverse_km2": 3.0}
        out = Path(self.tmpdir) / "test_report.json"
        save_ldn_planning_json(out, task_name="Test", arr_summary=arr)
        self.assertTrue(out.exists())
        data = json.loads(out.read_text())
        self.assertEqual(data["task_name"], "Test")
        self.assertIn("arr", data)

    def test_scenario_breakdown_sheets_written(self):
        from te_algorithms.gdal.land_deg.ldn_planning_report import (
            save_ldn_planning_excel,
        )
        from pathlib import Path

        scenario_summary = {
            "gains_km2_reverse": 3.0,
            "avoided_losses_km2_reduce": 2.0,
            "avoided_losses_km2_avoid": 1.0,
            "total_gains_km2": 3.0,
            "total_avoided_losses_km2": 3.0,
            "net_balance_note": "test",
            "per_target": [],
            "by_land_type": [
                {
                    "name": "Cropland",
                    "code": 3,
                    "gains_km2": 2.0,
                    "avoided_losses_reduce_km2": 1.0,
                    "avoided_losses_avoid_km2": 0.5,
                    "total_gains_km2": 2.0,
                    "total_avoided_losses_km2": 1.5,
                },
            ],
            "by_zone": [
                {
                    "name": "Region A",
                    "code": 1,
                    "gains_km2": 1.0,
                    "avoided_losses_reduce_km2": 1.0,
                    "avoided_losses_avoid_km2": 0.5,
                    "total_gains_km2": 1.0,
                    "total_avoided_losses_km2": 1.5,
                },
            ],
        }
        out = Path(self.tmpdir) / "test_scenario_report.xlsx"
        save_ldn_planning_excel(out, scenario_summary=scenario_summary)
        self.assertTrue(out.exists())

        import openpyxl

        wb = openpyxl.load_workbook(out)
        self.assertIn("Scenario by Land Type", wb.sheetnames)
        self.assertIn("Scenario by Jurisdiction", wb.sheetnames)


class TestBauScenarioComparison(unittest.TestCase):
    """Pure-Python tests for project_scenario_against_bau (no GDAL needed)."""

    def _bau(self):
        return {
            "year_initial": 2000,
            "year_final": 2015,
            "target_year": 2030,
            "degraded_area_baseline_km2": 100.0,
            "degraded_area_reporting_km2": 120.0,
            "bau_projection_2030_km2": 150.0,
            "ldntarget_km2": 100.0,
            "shortfall_km2": 50.0,
            "zones": [],
        }

    def test_no_reporting_returns_none(self):
        from te_algorithms.gdal.land_deg.ldn_planning import (
            project_scenario_against_bau,
        )

        bau = self._bau()
        bau["bau_projection_2030_km2"] = None
        result = project_scenario_against_bau(bau, {}, 2030)
        self.assertIsNone(result)

    def test_partial_gap_close(self):
        from te_algorithms.gdal.land_deg.ldn_planning import (
            project_scenario_against_bau,
        )

        scenario = {"total_gains_km2": 10.0, "total_avoided_losses_km2": 20.0}
        p = project_scenario_against_bau(self._bau(), scenario, 2030)
        # BAU projection 150, contribution 30 -> adjusted 120, target 100
        self.assertAlmostEqual(p["scenario_degraded_km2"], 120.0)
        self.assertAlmostEqual(p["bau_shortfall_km2"], 50.0)
        self.assertAlmostEqual(p["remaining_shortfall_km2"], 20.0)
        self.assertAlmostEqual(p["gap_closed_pct"], 60.0)
        self.assertFalse(p["neutral"])

    def test_neutrality_achieved(self):
        from te_algorithms.gdal.land_deg.ldn_planning import (
            project_scenario_against_bau,
        )

        scenario = {"total_gains_km2": 30.0, "total_avoided_losses_km2": 30.0}
        p = project_scenario_against_bau(self._bau(), scenario, 2030)
        # contribution 60 >= shortfall 50 -> adjusted floored at target
        self.assertTrue(p["neutral"])
        self.assertAlmostEqual(p["gap_closed_pct"], 100.0)


class TestZoneCombination(unittest.TestCase):
    """Tests for the multi-raster zone combination logic."""

    def test_combine_arrays_unique_combinations(self):
        from te_algorithms.gdal.land_deg import zones

        a = np.array([[1, 1], [2, 2]], dtype=np.int32)
        b = np.array([[1, 2], [1, 2]], dtype=np.int32)
        zone_ids, labels = zones.combine_arrays_to_zones([a, b], [-32768, -32768])
        # 4 unique combinations -> 4 sequential zones
        self.assertEqual(len(labels), 4)
        self.assertEqual(sorted(set(zone_ids.ravel())), [1, 2, 3, 4])
        # labels decode to the source combination "a_b"
        self.assertIn("1_1", labels.values())
        self.assertIn("2_2", labels.values())

    def test_combine_arrays_nodata_becomes_zero(self):
        from te_algorithms.gdal.land_deg import zones

        a = np.array([[1, -32768]], dtype=np.int32)
        b = np.array([[5, 5]], dtype=np.int32)
        zone_ids, labels = zones.combine_arrays_to_zones([a, b], [-32768, -32768])
        # The nodata pixel -> zone id 0 (outside)
        self.assertEqual(zone_ids[0, 1], 0)
        self.assertEqual(zone_ids[0, 0], 1)
        self.assertEqual(len(labels), 1)


if __name__ == "__main__":
    unittest.main()
