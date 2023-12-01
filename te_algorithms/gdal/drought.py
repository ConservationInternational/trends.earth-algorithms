import dataclasses
import datetime as dt
import json
import logging
import math
import multiprocessing
import tempfile
from copy import deepcopy
from pathlib import Path
from typing import Callable
from typing import Dict
from typing import List
from typing import Optional
from typing import Tuple

import marshmallow_dataclass
import numpy as np
import openpyxl
from osgeo import gdal
from te_schemas import reporting
from te_schemas import SchemaBase
from te_schemas import schemas
from te_schemas.aoi import AOI
from te_schemas.datafile import DataFile
from te_schemas.jobs import Job
from te_schemas.results import Band
from te_schemas.results import DataType
from te_schemas.results import Raster
from te_schemas.results import RasterFileType
from te_schemas.results import RasterResults
from te_schemas.results import URI

from . import util
from . import workers
from . import xl
from .. import __release_date__
from .. import __version__
from .drought_numba import *
from .util_numba import *

NODATA_VALUE = -32768
MASK_VALUE = -32767

POPULATION_BAND_NAME = "Population (number of people)"
SPI_BAND_NAME = "Standardized Precipitation Index (SPI)"
JRC_BAND_NAME = "Drought Vulnerability (JRC)"
WATER_MASK_BAND_NAME = "Water mask"
SPI_MIN_OVER_PERIOD_BAND_NAME = "Minimum SPI over period"
POP_AT_SPI_MIN_OVER_PERIOD_BAND_NAME = "Population at minimum SPI over period"

logger = logging.getLogger(__name__)


@marshmallow_dataclass.dataclass
class SummaryTableDrought(SchemaBase):
    annual_area_by_drought_class: List[Dict[int, float]]
    annual_population_by_drought_class_total: List[Dict[int, float]]
    annual_population_by_drought_class_male: List[Dict[int, float]]
    annual_population_by_drought_class_female: List[Dict[int, float]]
    dvi_value_sum_and_count: Tuple[float, int]

    def cast_to_cpython(self):
        self.annual_area_by_drought_class = cast_numba_int_dict_list_to_cpython(
            self.annual_area_by_drought_class
        )
        self.annual_population_by_drought_class_total = (
            cast_numba_int_dict_list_to_cpython(
                self.annual_population_by_drought_class_total
            )
        )
        self.annual_population_by_drought_class_male = (
            cast_numba_int_dict_list_to_cpython(
                self.annual_population_by_drought_class_male
            )
        )
        self.annual_population_by_drought_class_female = (
            cast_numba_int_dict_list_to_cpython(
                self.annual_population_by_drought_class_female
            )
        )


def _accumulate_drought_summary_tables(
    tables: List[SummaryTableDrought],
) -> SummaryTableDrought:
    if len(tables) == 1:
        return tables[0]
    else:
        out = tables[0]

        for table in tables[1:]:
            out.annual_area_by_drought_class = [
                util.accumulate_dicts([a, b])
                for a, b in zip(
                    out.annual_area_by_drought_class, table.annual_area_by_drought_class
                )
            ]
            out.annual_population_by_drought_class_total = [
                util.accumulate_dicts([a, b])
                for a, b in zip(
                    out.annual_population_by_drought_class_total,
                    table.annual_population_by_drought_class_total,
                )
            ]
            out.annual_population_by_drought_class_male = [
                util.accumulate_dicts([a, b])
                for a, b in zip(
                    out.annual_population_by_drought_class_male,
                    table.annual_population_by_drought_class_male,
                )
            ]
            out.annual_population_by_drought_class_female = [
                util.accumulate_dicts([a, b])
                for a, b in zip(
                    out.annual_population_by_drought_class_female,
                    table.annual_population_by_drought_class_female,
                )
            ]
            out.dvi_value_sum_and_count = (
                out.dvi_value_sum_and_count[0] + table.dvi_value_sum_and_count[0],
                out.dvi_value_sum_and_count[1] + table.dvi_value_sum_and_count[1],
            )

        return out


@dataclasses.dataclass()
class DroughtSummaryParams(SchemaBase):
    in_df: DataFile
    out_file: str
    drought_period: int
    mask_file: str


def _expand_dims(array, in_array):
    """
    Avoid indexing errors when a tile winds up with a single row or col

    Need to add a dimension to arrays if the first or second dimension of
    in_array is 1 (meaning there is only 1 row or 1 column in a block) as
    otherwise the squeeze done after finding the max drought indices will
    result in a missing dimension of the max drought arrays, throwing off the
    later masking code as all those arrays will have full dimensions
    """

    for axis in [1, 2]:
        if in_array.shape[axis] == 1:
            logger.debug(
                "expanding dim %s as in_array.shape is %s", axis - 1, in_array.shape
            )
            array = np.expand_dims(array, axis=axis - 1)

    return array


def _process_block(
    params: DroughtSummaryParams, in_array, mask, xoff: int, yoff: int, cell_areas_raw
) -> Tuple[SummaryTableDrought, Dict]:
    logger.debug("in_array.shape is %s", in_array.shape)

    write_arrays = {}

    # Make an array of the same size as the input arrays containing
    # the area of each cell (which is identical for all cells in a
    # given row - cell areas only vary among rows)
    cell_areas = np.repeat(cell_areas_raw, mask.shape[1], axis=1).astype(np.float64)

    spi_rows = params.in_df.indices_for_name(SPI_BAND_NAME)
    pop_rows_total = params.in_df.indices_for_name(
        POPULATION_BAND_NAME, field="type", field_filter="total"
    )
    pop_rows_male = params.in_df.indices_for_name(
        POPULATION_BAND_NAME, field="type", field_filter="male"
    )
    pop_rows_female = params.in_df.indices_for_name(
        POPULATION_BAND_NAME, field="type", field_filter="female"
    )

    try:
        water_mask_index = params.in_df.index_for_name(WATER_MASK_BAND_NAME)
    except IndexError:
        mask_water = False
    else:
        mask_water = True
        a_water_mask = in_array[water_mask_index, :, :]

    # There should either be one pop row for each SPI row (if using total pop)
    # or two rows per SPI row (if using gender disaggregated population data)

    assert len(spi_rows) == len(pop_rows_total) or len(spi_rows) * 2 == (
        len(pop_rows_male) + len(pop_rows_female)
    )

    if len(pop_rows_male) >= 1:
        pop_by_sex = True
    else:
        pop_by_sex = False

    # Calculate annual totals of area and population exposed to drought
    annual_area_by_drought_class = []
    annual_population_by_drought_class_total = []
    annual_population_by_drought_class_male = []
    annual_population_by_drought_class_female = []

    for row_num in range(len(spi_rows)):
        spi_row = spi_rows[row_num]
        a_drought_class = drought_class(in_array[spi_row, :, :])

        annual_area_by_drought_class.append(
            zonal_total(a_drought_class, cell_areas, mask)
        )

        if pop_by_sex:
            pop_row_male = pop_rows_male[row_num]
            pop_row_female = pop_rows_female[row_num]

            a_pop_male = in_array[pop_row_male, :, :]
            a_pop_male_recoded = a_pop_male.copy().astype(np.float64)
            a_pop_male_recoded[a_pop_male == NODATA_VALUE] = 0

            if mask_water:
                a_pop_male_recoded[a_water_mask == 1] = 0
            annual_population_by_drought_class_male.append(
                zonal_total(a_drought_class, a_pop_male_recoded, mask)
            )

            a_pop_female = in_array[pop_row_female, :, :]
            a_pop_female_recoded = a_pop_female.copy().astype(np.float64)
            a_pop_female_recoded[a_pop_female == NODATA_VALUE] = 0

            if mask_water:
                a_pop_female_recoded[a_water_mask == 1] = 0
            annual_population_by_drought_class_female.append(
                zonal_total(a_drought_class, a_pop_female_recoded, mask)
            )

            a_pop_total_recoded = a_pop_male_recoded + a_pop_female_recoded

        else:
            pop_row_total = pop_rows_total[row_num]
            a_pop_total = in_array[pop_row_total, :, :]
            a_pop_total_recoded = a_pop_total.copy().astype(np.float64)
            a_pop_total_recoded[a_pop_total == NODATA_VALUE] = 0

            if mask_water:
                a_pop_total_recoded[a_water_mask == 1] = 0

        annual_population_by_drought_class_total.append(
            zonal_total(a_drought_class, a_pop_total_recoded, mask)
        )

    # Calculate minimum SPI in blocks of length (in years) defined by
    # params.drought_period, and save the spi at that point as well as
    # population that was exposed to it at that point

    first_rows = [*range(0, len(spi_rows), params.drought_period)]

    for period_number, first_row in enumerate(first_rows):
        if (first_row + params.drought_period - 1) > len(spi_rows):
            last_row = len(spi_rows)
        else:
            last_row = first_row + params.drought_period - 1

        spis = in_array[spi_rows[first_row:last_row], :, :]

        # Max drought is at minimum SPI
        min_indices = np.expand_dims(np.argmin(spis, axis=0), axis=0)
        # squeeze to remove zero dim of len 1
        max_drought = np.take_along_axis(spis, min_indices, axis=0).squeeze()

        if pop_by_sex:
            pop_male = (
                in_array[pop_rows_male[first_row:last_row], :, :]
                .copy()
                .astype(np.float64)
            )
            pop_female = (
                in_array[pop_rows_female[first_row:last_row], :, :]
                .copy()
                .astype(np.float64)
            )
            pop_total = pop_male + pop_female
        else:
            pop_total = (
                in_array[pop_rows_total[first_row:last_row], :, :]
                .copy()
                .astype(np.float64)
            )
        pop_total_max_drought = np.take_along_axis(
            pop_total, min_indices, axis=0
        ).squeeze()

        if pop_by_sex:
            pop_female_max_drought = np.take_along_axis(
                pop_female, min_indices, axis=0
            ).squeeze()
            pop_male_max_drought = np.take_along_axis(
                pop_male, min_indices, axis=0
            ).squeeze()

        # Need to add a dimension to max_drought and pop_total_max_drought if
        # the second dimension of in_array is 1 (meaning there only 1 row in
        # this block) as otherwise after the squeeze we will be missing the
        # second dimension of pop_total_max_drought, throwing off the masking
        # and multiplication by cell_areas as all those arrays will have a
        # second dimension of 1

        pop_total_max_drought = _expand_dims(pop_total_max_drought, in_array)
        max_drought = _expand_dims(max_drought, in_array)

        if pop_by_sex:
            pop_female_max_drought = _expand_dims(pop_female_max_drought, in_array)
            pop_male_max_drought = _expand_dims(pop_male_max_drought, in_array)

        pop_total_max_drought[pop_total_max_drought == NODATA_VALUE] = 0
        pop_total_max_drought[max_drought < -1000] = -pop_total_max_drought[
            max_drought < -1000
        ]
        # Set water to NODATA_VALUE as requested by UNCCD for Prais

        logger.debug("pop_total_max_drought.shape %s", pop_total_max_drought.shape)

        if mask_water:
            logger.debug("a_water_mask.shape %s", a_water_mask.shape)
            pop_total_max_drought[a_water_mask == 1] = NODATA_VALUE

        # Add one as output band numbers start at 1, not zero
        write_arrays[2 * period_number + 1] = {
            "array": max_drought,
            "xoff": xoff,
            "yoff": yoff,
        }

        # Add two as output band numbers start at 1, not zero, and this is the
        # second band for this period
        write_arrays[2 * period_number + 2] = {
            "array": pop_total_max_drought,
            "xoff": xoff,
            "yoff": yoff,
        }

        if pop_by_sex and period_number == (len(first_rows) - 1):
            # Only write out population disaggregated by sex for the last
            # period
            pop_female_max_drought[pop_female_max_drought == NODATA_VALUE] = 0
            pop_female_max_drought[max_drought < -1000] = -pop_female_max_drought[
                max_drought < -1000
            ]
            # Set water to NODATA_VALUE as requested by UNCCD for Prais

            if mask_water:
                pop_female_max_drought[a_water_mask == 1] = NODATA_VALUE

            # Add two as output band numbers start at 1, not zero, and this is
            # the third band for this period
            write_arrays[2 * period_number + 3] = {
                "array": pop_female_max_drought,
                "xoff": xoff,
                "yoff": yoff,
            }

            pop_male_max_drought[pop_male_max_drought == NODATA_VALUE] = 0
            pop_male_max_drought[max_drought < -1000] = -pop_male_max_drought[
                max_drought < -1000
            ]
            # Set water to NODATA_VALUE as requested by UNCCD for Prais

            if mask_water:
                pop_male_max_drought[a_water_mask == 1] = NODATA_VALUE

            # Add two as output band numbers start at 1, not zero, and this is
            # the fourth band for this period
            write_arrays[2 * period_number + 4] = {
                "array": pop_male_max_drought,
                "xoff": xoff,
                "yoff": yoff,
            }

    jrc_row = params.in_df.index_for_name(JRC_BAND_NAME)
    dvi_value_sum_and_count = jrc_sum_and_count(in_array[jrc_row, :, :], mask)

    out_table = SummaryTableDrought(
        annual_area_by_drought_class,
        annual_population_by_drought_class_total,
        annual_population_by_drought_class_male,
        annual_population_by_drought_class_female,
        dvi_value_sum_and_count,
    )

    return out_table, write_arrays


@dataclasses.dataclass()
class LineParams:
    params: DroughtSummaryParams
    image_info: util.ImageInfo
    y: int
    win_ysize: int
    lat: float


def _get_cell_areas(y, lat, win_y_size, image_info):
    cell_areas = (
        np.array(
            [
                calc_cell_area(
                    lat + image_info.pixel_height * n,
                    lat + image_info.pixel_height * (n + 1),
                    image_info.pixel_width,
                )
                for n in range(win_y_size)
            ]
        )
        * 1e-6
    )  # 1e-6 is to convert from meters to kilometers
    cell_areas.shape = (cell_areas.size, 1)

    return cell_areas


def _process_line(line_params: LineParams):
    mask_ds = gdal.Open(line_params.params.mask_file)
    mask_band = mask_ds.GetRasterBand(1)
    src_ds = gdal.Open(str(line_params.params.in_df.path))

    cell_areas = _get_cell_areas(
        line_params.y, line_params.lat, line_params.win_ysize, line_params.image_info
    )

    results = []

    for x in range(
        0, line_params.image_info.x_size, line_params.image_info.x_block_size
    ):
        if x + line_params.image_info.x_block_size < line_params.image_info.x_size:
            win_xsize = line_params.image_info.x_block_size
        else:
            win_xsize = line_params.image_info.x_size - x

        logger.debug("image_info: %s", line_params.image_info)
        logger.debug("x %s, win_xsize %s", x, win_xsize)

        src_array = src_ds.ReadAsArray(
            xoff=x, yoff=line_params.y, xsize=win_xsize, ysize=line_params.win_ysize
        )

        mask_array = mask_band.ReadAsArray(
            xoff=x,
            yoff=line_params.y,
            win_xsize=win_xsize,
            win_ysize=line_params.win_ysize,
        )
        mask_array = mask_array == MASK_VALUE

        result = _process_block(
            line_params.params, src_array, mask_array, x, line_params.y, cell_areas
        )

        results.append(result)

    return results


def _get_n_pop_band_for_type(dfs, pop_type):
    n_bands = 0

    for df in dfs:
        n_bands += len(
            df.indices_for_name(
                POPULATION_BAND_NAME, field="type", field_filter=pop_type
            )
        )

    return n_bands


def _get_n_spi_bands(dfs):
    n_bands = 0

    for df in dfs:
        n_bands += len(df.indices_for_name(SPI_BAND_NAME))

    return n_bands


def _have_pop_by_sex(in_dfs):
    n_spi_bands = _get_n_spi_bands(in_dfs)
    n_total_pop_bands = _get_n_pop_band_for_type(in_dfs, "total")
    n_female_pop_bands = _get_n_pop_band_for_type(in_dfs, "female")
    n_male_pop_bands = _get_n_pop_band_for_type(in_dfs, "male")

    logger.debug(
        "n_total_pop_bands %s, n_female_pop_bands %s, n_male_pop_bands %s",
        n_total_pop_bands,
        n_female_pop_bands,
        n_male_pop_bands,
    )

    assert n_spi_bands == n_total_pop_bands or n_spi_bands * 2 == (
        n_male_pop_bands + n_female_pop_bands
    )

    if n_male_pop_bands >= 1:
        return True
    else:
        return False


class DroughtSummary:
    def __init__(self, params: DroughtSummaryParams):
        self.params = params
        self.image_info = util.get_image_info(self.params.in_df.path)

    def is_killed(self):
        return False

    def emit_progress(self, *args):
        """Reimplement to display progress messages"""
        util.log_progress(
            *args, message=f"Processing drought summary for {self.params.in_df.path}"
        )

    def get_line_params(self):
        """Make a list of parameters to use in the _process_line function"""
        # Set initial lat to the top left corner latitude
        src_ds = gdal.Open(str(self.params.in_df.path))
        src_gt = src_ds.GetGeoTransform()
        lat = src_gt[3]

        logger.debug(
            "getting line params for image with xsize " "%s, and ysize %s",
            src_ds.RasterXSize,
            src_ds.RasterYSize,
        )

        line_params = []

        for y in range(0, self.image_info.y_size, self.image_info.y_block_size):
            if y + self.image_info.y_block_size < self.image_info.y_size:
                win_ysize = self.image_info.y_block_size
            else:
                win_ysize = self.image_info.y_size - y

            line_params.append(
                LineParams(self.params, self.image_info, y, win_ysize, lat)
            )

            lat += self.image_info.pixel_height * win_ysize

        return line_params

    def process_lines(self, line_params_list):
        out_ds = self._get_out_ds()

        out = []

        for n, line_params in enumerate(line_params_list):
            self.emit_progress(n / len(line_params_list))
            results = _process_line(line_params)

            for result in results:
                out.append(result[0])

                for key, value in result[1].items():
                    out_ds.GetRasterBand(key).WriteArray(**value)

        out = _accumulate_drought_summary_tables(out)

        return out

    def _get_out_ds(self):
        n_out_bands = int(
            2
            * math.ceil(
                len(self.params.in_df.indices_for_name(SPI_BAND_NAME))
                / self.params.drought_period
            )
        )

        if _have_pop_by_sex([self.params.in_df]):
            # If have population disaggregated by sex, then the total
            # population at max drought layer is written for each period except
            # for the last, which also includes male/female totals - so need
            # two more out bands
            n_out_bands += 2

        out_ds = util.setup_output_image(
            self.params.in_df.path, self.params.out_file, n_out_bands, self.image_info
        )

        return out_ds


def _get_population_band_instance(population_type, year_initial, year_final):
    return Band(
        name=POP_AT_SPI_MIN_OVER_PERIOD_BAND_NAME,
        no_data_value=NODATA_VALUE,
        metadata={
            "year_initial": year_initial,
            "year_final": year_final,
            "type": population_type,
        },
        activated=True,
    )


def summarise_drought_vulnerability(
    drought_job: Job,
    aoi: AOI,
    job_output_path: Path,
    n_cpus: int = multiprocessing.cpu_count() - 1,
) -> Job:
    logger.debug("at top of summarise_drought_vulnerability")

    params = drought_job.params

    drought_period = 4

    spi_dfs = _prepare_dfs(
        params["layer_spi_path"],
        params["layer_spi_bands"],
        params["layer_spi_band_indices"],
    )

    population_dfs = _prepare_dfs(
        params["layer_population_path"],
        params["layer_population_bands"],
        params["layer_population_band_indices"],
    )

    jrc_df = _prepare_dfs(
        params["layer_jrc_path"],
        [params["layer_jrc_band"]],
        [params["layer_jrc_band_index"]],
    )

    if params.get("layer_water_path") is not None:
        # Water layers are optional - if not provided then water won't be
        # masked
        water_df = _prepare_dfs(
            params["layer_water_path"],
            [params["layer_water_band"]],
            [params["layer_water_band_index"]],
        )
    else:
        water_df = []

    summary_table, out_path = _compute_drought_summary_table(
        aoi=aoi,
        compute_bbs_from=params["layer_spi_path"],
        output_job_path=job_output_path.parent / f"{job_output_path.stem}.json",
        in_dfs=spi_dfs + population_dfs + jrc_df + water_df,
        drought_period=drought_period,
        n_cpus=n_cpus,
    )

    out_bands = []
    logger.info(
        f"Processing for years {params['layer_spi_years'][0]} - "
        f"{int(params['layer_spi_years'][-1])}"
    )

    year_initials = [
        *range(
            int(params["layer_spi_years"][0]),
            int(params["layer_spi_years"][-1]),
            drought_period,
        )
    ]

    for period_number, year_initial in enumerate(year_initials):
        if (year_initial + drought_period - 1) > params["layer_spi_years"][-1]:
            year_final = params["layer_spi_years"][-1]
        else:
            year_final = year_initial + drought_period - 1

        out_bands.append(
            Band(
                name=SPI_MIN_OVER_PERIOD_BAND_NAME,
                no_data_value=NODATA_VALUE,
                metadata={
                    "year_initial": year_initial,
                    "year_final": year_final,
                    "lag": int(params["layer_spi_lag"]),
                },
                activated=True,
            )
        )

        out_bands.append(
            _get_population_band_instance("total", year_initial, year_final)
        )

        if _have_pop_by_sex(spi_dfs + population_dfs) and period_number == (
            len(year_initials) - 1
        ):
            out_bands.append(
                _get_population_band_instance("female", year_initial, year_final)
            )
            out_bands.append(
                _get_population_band_instance("male", year_initial, year_final)
            )

    out_df = DataFile(out_path.name, out_bands)

    # Also save bands to a key file for ease of use in PRAIS
    key_json = job_output_path.parent / f"{job_output_path.stem}_band_key.json"
    with open(key_json, "w") as f:
        json.dump(DataFile.Schema().dump(out_df), f, indent=4)

    summary_json_output_path = (
        job_output_path.parent / f"{job_output_path.stem}_summary.json"
    )
    report_json = save_reporting_json(
        summary_json_output_path,
        summary_table,
        drought_job.params,
        drought_job.task_name,
        aoi,
    )

    summary_table_output_path = (
        job_output_path.parent / f"{job_output_path.stem}_summary.xlsx"
    )
    save_summary_table_excel(
        summary_table_output_path,
        summary_table,
        years=[int(y) for y in params["layer_spi_years"]],
    )

    results = RasterResults(
        name="drought_vulnerability_summary",
        uri=URI(uri=out_path),
        rasters={
            DataType.INT16.value: Raster(
                uri=URI(uri=out_path),
                bands=out_df.bands,
                datatype=DataType.INT16,
                filetype=RasterFileType.COG,
            )
        },
        data={"report": report_json},
    )

    return results


def _prepare_dfs(path, band_str_list, band_indices) -> List[DataFile]:
    dfs = []

    for band_str, band_index in zip(band_str_list, band_indices):
        band = Band(**band_str)
        dfs.append(DataFile(path=util.save_vrt(path, band_index), bands=[band]))

    return dfs


def _aoi_process_multiprocess(inputs, n_cpus):
    with multiprocessing.get_context("spawn").Pool(n_cpus) as p:
        n = 0

        results = []

        for output in p.imap_unordered(_summarize_tile, inputs):
            util.log_progress(
                n / len(inputs), message="Processing drought summaries overall progress"
            )
            error_message = output[1]

            if error_message is not None:
                p.terminate()

                break

            results.append(output[0])
            n += 1

    return results


def _aoi_process_sequential(inputs):
    results = []

    for item in inputs:
        n = 0
        output = _summarize_tile(item)
        util.log_progress(
            n / len(inputs), message="Processing drought summaries overall progress"
        )
        error_message = output[1]

        if error_message is not None:
            break

        results.append(output[0])
        n += 1

    return results


def _summarize_over_aoi(
    wkt_aoi,
    pixel_aligned_bbox,
    in_dfs: List[DataFile],
    output_tif_path: Path,
    mask_worker_process_name,
    drought_worker_process_name,
    drought_period: int,
    n_cpus: int,
    translate_worker_function: Callable = None,
    translate_worker_params: dict = None,
    mask_worker_function: Callable = None,
    mask_worker_params: dict = None,
    drought_worker_function: Callable = None,
    drought_worker_params: dict = None,
) -> Tuple[Optional[SummaryTableDrought], str]:
    # Combine all raster into a VRT and crop to the AOI
    indic_vrt = tempfile.NamedTemporaryFile(
        suffix="_drought_indicators.vrt", delete=False
    ).name
    indic_vrt = Path(indic_vrt)
    logger.info("Saving indicator VRT to {}".format(indic_vrt))
    # The plus one is because band numbers start at 1, not zero
    gdal.BuildVRT(
        str(indic_vrt),
        [item.path for item in in_dfs],
        outputBounds=pixel_aligned_bbox,
        resolution="highest",
        resampleAlg=gdal.GRA_NearestNeighbour,
        separate=True,
    )

    indic_reproj = tempfile.NamedTemporaryFile(
        suffix="_drought_indicators_reproj_tiles.tif", delete=False
    ).name
    indic_reproj = Path(indic_reproj)
    logger.info(f"Reprojecting inputs and saving to {indic_reproj}")

    error_message = ""

    if translate_worker_function:
        tiles = translate_worker_function(
            indic_vrt, str(indic_reproj), **translate_worker_params
        )

    else:
        translate_worker = workers.CutTiles(
            indic_vrt, n_cpus, indic_reproj, gdal.GDT_Int32
        )
        tiles = translate_worker.work()
        logger.debug("Tiles are %s", tiles)

    if tiles:
        out_files = [
            output_tif_path.parent / (output_tif_path.stem + f"_{n}.tif")
            for n in range(len(tiles))
        ]
        inputs = [
            SummarizeTileInputs(
                tile=tile,
                out_file=out_file,
                aoi=wkt_aoi,
                drought_period=drought_period,
                in_dfs=in_dfs,
                mask_worker_function=mask_worker_function,
                mask_worker_params=mask_worker_params,
                drought_worker_function=drought_worker_function,
                drought_worker_params=drought_worker_params,
            )
            for tile, out_file in zip(tiles, out_files)
        ]

        if n_cpus > 1:
            results = _aoi_process_multiprocess(inputs, n_cpus)
        else:
            results = _aoi_process_sequential(inputs)

        results = _accumulate_drought_summary_tables(results)
    else:
        error_message = "Error reprojecting layers."
        results = None

    return results, out_files, error_message


@dataclasses.dataclass()
class SummarizeTileInputs:
    tile: Path
    out_file: Path
    aoi: str
    drought_period: int
    in_dfs: List[DataFile]
    mask_worker_function: Callable = None
    mask_worker_params: dict = None
    drought_worker_function: Callable = None
    drought_worker_params: dict = None


def _summarize_tile(tile_input):
    logger.info("Processing tile %s", tile_input.tile)
    # Compute a mask layer that will be used in the tabulation code to
    # mask out areas outside of the AOI. Do this instead of using
    # gdal.Clip to save having to clip and rewrite all of the layers in
    # the VRT
    mask_tif = tempfile.NamedTemporaryFile(
        suffix="_drought_mask.tif", delete=False
    ).name

    logger.info(f"Saving mask to {mask_tif}")
    geojson = util.wkt_geom_to_geojson_file_string(tile_input.aoi)

    error_message = None

    if tile_input.mask_worker_function:
        mask_result = tile_input.mask_worker_function(
            mask_tif, geojson, str(tile_input.tile), **tile_input.mask_worker_params
        )
    else:
        mask_worker = workers.Mask(mask_tif, geojson, str(tile_input.tile))
        mask_result = mask_worker.work()

    if mask_result:
        # Combine all in_dfs together and update path to refer to indicator
        # VRT
        in_df = DataFile(
            tile_input.tile, [b for d in tile_input.in_dfs for b in d.bands]
        )
        params = DroughtSummaryParams(
            in_df=in_df,
            out_file=str(tile_input.out_file),
            mask_file=mask_tif,
            drought_period=tile_input.drought_period,
        )

        logger.info(
            "Calculating summary table and saving " f"rasters to {tile_input.out_file}"
        )

        if tile_input.drought_worker_function:
            result = tile_input.drought_worker_function(
                params, **tile_input.drought_worker_params
            )
        else:
            summarizer = DroughtSummary(params)
            result = summarizer.process_lines(summarizer.get_line_params())

        if not result:
            if result.is_killed():
                error_message = (
                    "Cancelled calculation of summary " f"table for {tile_input.tile}."
                )
            else:
                error_message = (
                    f"Error calculating summary table for  {tile_input.tile}."
                )
                result = None
        else:
            result.cast_to_cpython()
    else:
        error_message = f"Error creating mask for tile {tile_input.tile}."
        result = None

    return result, error_message


def _compute_drought_summary_table(
    aoi,
    compute_bbs_from,
    in_dfs,
    output_job_path: Path,
    drought_period: int,
    n_cpus: int,
) -> Tuple[SummaryTableDrought, Path, Path]:
    """Computes summary table and the output tif file(s)"""
    wkt_aois = aoi.meridian_split(as_extent=False, out_format="wkt")
    bbs = aoi.get_aligned_output_bounds(compute_bbs_from)
    assert len(wkt_aois) == len(bbs)

    output_name_pattern = {
        1: f"{output_job_path.stem}" + ".tif",
        2: f"{output_job_path.stem}" + "_{index}.tif",
    }[len(wkt_aois)]
    mask_name_fragment = {
        1: "Generating mask",
        2: "Generating mask (part {index} of 2)",
    }[len(wkt_aois)]
    drought_name_fragment = {
        1: "Calculating summary table",
        2: "Calculating summary table (part {index} of 2)",
    }[len(wkt_aois)]

    summary_tables = []
    out_paths = []

    for index, (wkt_aoi, pixel_aligned_bbox) in enumerate(zip(wkt_aois, bbs), start=1):
        logger.info(f"Calculating summary table {index} of {len(wkt_aois)}")
        out_path = output_job_path.parent / output_name_pattern.format(
            index=index
        )  # This is a base path - there may be multiple tiles generated below
        result, out_files, error_message = _summarize_over_aoi(
            wkt_aoi=wkt_aoi,
            pixel_aligned_bbox=pixel_aligned_bbox,
            output_tif_path=out_path,
            mask_worker_process_name=mask_name_fragment.format(index=index),
            drought_worker_process_name=drought_name_fragment.format(index=index),
            in_dfs=deepcopy(in_dfs),
            drought_period=drought_period,
            n_cpus=n_cpus,
        )
        out_paths.extend(out_files)

        if result is None:
            raise RuntimeError(error_message)
        else:
            summary_tables.append(result)

    summary_table = _accumulate_drought_summary_tables(summary_tables)

    if len(out_paths) > 1:
        out_path = output_job_path.parent / f"{output_job_path.stem}.vrt"
        gdal.BuildVRT(str(out_path), [str(p) for p in out_paths])
    else:
        out_path = out_paths[0]

    return summary_table, out_path


def save_summary_table_excel(
    output_path: Path, summary_table: SummaryTableDrought, years: List[int]
):
    """Save summary table into an xlsx file on disk"""
    template_summary_table_path = (
        Path(__file__).parents[1] / "data/summary_table_drought.xlsx"
    )
    workbook = openpyxl.load_workbook(str(template_summary_table_path))
    _render_drought_workbook(workbook, summary_table, years)
    try:
        workbook.save(output_path)
        logger.info("Indicator table saved to {}".format(output_path))

    except OSError:
        error_message = (
            f"Error saving output table - check that {output_path!r} is accessible "
            f"and not already open."
        )
        logger.error(error_message)


def _get_population_list_by_drought_class(pop_by_drought, pop_type):
    return reporting.PopulationList(
        "Population by drought class",
        [
            reporting.Population(
                "Mild drought", pop_by_drought.get(1, 0), type=pop_type
            ),
            reporting.Population(
                "Moderate drought", pop_by_drought.get(2, 0), type=pop_type
            ),
            reporting.Population(
                "Severe drought", pop_by_drought.get(3, 0), type=pop_type
            ),
            reporting.Population(
                "Extreme drought", pop_by_drought.get(4, 0), type=pop_type
            ),
            reporting.Population(
                "Non-drought", pop_by_drought.get(0, 0), type=pop_type
            ),
            reporting.Population(
                "No data", pop_by_drought.get(-32768, 0), type=pop_type
            ),
        ],
    )


def save_reporting_json(
    output_path: Path, st: SummaryTableDrought, params: dict, task_name: str, aoi: AOI
):
    drought_tier_one = {}
    drought_tier_two = {}

    for n, year in enumerate(
        range(int(params["layer_spi_years"][0]), int(params["layer_spi_years"][-1]) + 1)
    ):
        total_land_area = sum(
            [
                value
                for key, value in st.annual_area_by_drought_class[n].items()
                if key != MASK_VALUE
            ]
        )
        logger.debug(f"Total land area in {year} per drought data {total_land_area}")

        drought_tier_one[year] = reporting.AreaList(
            "Area by drought class",
            "sq km",
            [
                reporting.Area(
                    "Mild drought", st.annual_area_by_drought_class[n].get(1, 0.0)
                ),
                reporting.Area(
                    "Moderate drought", st.annual_area_by_drought_class[n].get(2, 0.0)
                ),
                reporting.Area(
                    "Severe drought", st.annual_area_by_drought_class[n].get(3, 0.0)
                ),
                reporting.Area(
                    "Extreme drought", st.annual_area_by_drought_class[n].get(4, 0.0)
                ),
                reporting.Area(
                    "Non-drought", st.annual_area_by_drought_class[n].get(0, 0.0)
                ),
                reporting.Area(
                    "No data", st.annual_area_by_drought_class[n].get(-32768, 0.0)
                ),
            ],
        )

        drought_tier_two[year] = {
            "Total population": _get_population_list_by_drought_class(
                st.annual_population_by_drought_class_total[n], "Total population"
            )
        }

        if len(st.annual_population_by_drought_class_male) > 0:
            drought_tier_two[year][
                "Male population"
            ] = _get_population_list_by_drought_class(
                st.annual_population_by_drought_class_male[n], "Male population"
            )

        if len(st.annual_population_by_drought_class_female) > 0:
            drought_tier_two[year][
                "Female population"
            ] = _get_population_list_by_drought_class(
                st.annual_population_by_drought_class_female[n], "Female population"
            )

    if st.dvi_value_sum_and_count[1] == 0:
        dvi_out = None
    else:
        dvi_out = st.dvi_value_sum_and_count[0] / st.dvi_value_sum_and_count[1]
    drought_tier_three = {2018: reporting.Value("Mean value", dvi_out)}

    ##########################################################################
    # Format final JSON output
    te_summary = reporting.TrendsEarthDroughtSummary(
        metadata=reporting.ReportMetadata(
            title="Trends.Earth Summary Report",
            date=dt.datetime.now(dt.timezone.utc),
            trends_earth_version=schemas.TrendsEarthVersion(
                version=__version__,
                revision=None,
                release_date=dt.datetime.strptime(
                    __release_date__, "%Y/%m/%d %H:%M:%SZ"
                ),
            ),
            area_of_interest=schemas.AreaOfInterest(
                name=task_name,  # TODO replace this with area of interest name once implemented in TE
                geojson=aoi.get_geojson(),
                crs_wkt=aoi.get_crs_wkt(),
            ),
        ),
        drought=reporting.DroughtReport(
            tier_one=drought_tier_one,
            tier_two=drought_tier_two,
            tier_three=drought_tier_three,
        ),
    )

    try:
        te_summary_json = json.loads(
            reporting.TrendsEarthDroughtSummary.Schema().dumps(te_summary)
        )
        with open(output_path, "w") as f:
            json.dump(te_summary_json, f, indent=4)

        return te_summary_json

    except OSError:
        logger.error("Error saving {output_path}")
        error_message = (
            "Error saving indicator table JSON - check that "
            f"{output_path} is accessible and not already open."
        )

        return None


def _render_drought_workbook(
    template_workbook, summary_table: SummaryTableDrought, years: List[int]
):
    _write_drought_area_sheet(
        template_workbook["Area under drought by year"], summary_table, years
    )

    _write_drought_pop_total_sheet(
        template_workbook["Pop under drought (total)"], summary_table, years
    )

    _write_dvi_sheet(
        template_workbook["Drought Vulnerability Index"], summary_table, years
    )

    return template_workbook


def _get_col_for_drought_class(annual_values_by_drought, drought_code):
    out = []

    for values_by_drought in annual_values_by_drought:
        out.append(values_by_drought.get(drought_code, 0))

    return np.array(out)


def _write_dvi_sheet(sheet, st: SummaryTableDrought, years):
    # Make this more informative when fuller DVI calculations are available...
    cell = sheet.cell(6, 2)
    cell.value = 2018
    cell = sheet.cell(6, 3)

    if st.dvi_value_sum_and_count[1] == 0:
        dvi_out = None
    else:
        dvi_out = st.dvi_value_sum_and_count[0] / st.dvi_value_sum_and_count[1]
    cell.value = dvi_out

    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet, "H1")


def _write_drought_area_sheet(sheet, st: SummaryTableDrought, years):
    xl.write_col_to_sheet(sheet, np.array(years), 2, 7)
    xl.write_col_to_sheet(
        sheet, _get_col_for_drought_class(st.annual_area_by_drought_class, 1), 4, 7
    )
    xl.write_col_to_sheet(
        sheet, _get_col_for_drought_class(st.annual_area_by_drought_class, 2), 6, 7
    )
    xl.write_col_to_sheet(
        sheet, _get_col_for_drought_class(st.annual_area_by_drought_class, 3), 8, 7
    )
    xl.write_col_to_sheet(
        sheet, _get_col_for_drought_class(st.annual_area_by_drought_class, 4), 10, 7
    )
    xl.write_col_to_sheet(
        sheet, _get_col_for_drought_class(st.annual_area_by_drought_class, 0), 12, 7
    )
    xl.write_col_to_sheet(
        sheet,
        _get_col_for_drought_class(st.annual_area_by_drought_class, -32768),
        14,
        7,
    )
    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet, "L1")


def _write_drought_pop_columns(
    sheet, drought_class_annual_totals: list, years, initial_row: int
):
    xl.write_col_to_sheet(sheet, np.array(years), 2, initial_row)
    xl.write_col_to_sheet(
        sheet,
        _get_col_for_drought_class(drought_class_annual_totals, 1),
        4,
        initial_row,
    )
    xl.write_col_to_sheet(
        sheet,
        _get_col_for_drought_class(drought_class_annual_totals, 2),
        6,
        initial_row,
    )
    xl.write_col_to_sheet(
        sheet,
        _get_col_for_drought_class(drought_class_annual_totals, 3),
        8,
        initial_row,
    )
    xl.write_col_to_sheet(
        sheet,
        _get_col_for_drought_class(drought_class_annual_totals, 4),
        10,
        initial_row,
    )
    xl.write_col_to_sheet(
        sheet,
        _get_col_for_drought_class(drought_class_annual_totals, 0),
        12,
        initial_row,
    )
    xl.write_col_to_sheet(
        sheet,
        _get_col_for_drought_class(drought_class_annual_totals, -32768),
        14,
        initial_row,
    )


def _write_drought_pop_total_sheet(sheet, st: SummaryTableDrought, years):
    if len(st.annual_population_by_drought_class_total) > 0:
        _write_drought_pop_columns(
            sheet, st.annual_population_by_drought_class_total, years, initial_row=7
        )

    if len(st.annual_population_by_drought_class_female) > 0:
        _write_drought_pop_columns(
            sheet, st.annual_population_by_drought_class_female, years, initial_row=35
        )

    if len(st.annual_population_by_drought_class_male) > 0:
        _write_drought_pop_columns(
            sheet, st.annual_population_by_drought_class_male, years, initial_row=63
        )

    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet, "L1")
