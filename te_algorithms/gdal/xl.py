from pathlib import Path

from numpy import ndarray
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles.borders import Border
from openpyxl.styles.borders import Side


def maybe_add_image_to_sheet(image_filename: str, sheet, place="H1"):
    try:
        from openpyxl.drawing.image import Image

        image_path = Path(__file__).parents[1] / "data" / image_filename
        logo = Image(image_path)
        sheet.add_image(logo, place)
    except ImportError:
        # add_image will fail on computers without PIL installed (this will be
        # an issue on some Macs, likely others). it is only used here to add
        # our logo, so no big deal.
        pass


thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def write_col_to_sheet(
    sheet,
    d,
    col,
    first_row,
    header=False,
    border=False,
    wrap=False,
    number_format="#,##0.00",
):
    if isinstance(d, ndarray):
        length = d.size
    else:
        length = len(d)
    for row in range(length):
        cell = sheet.cell(row=row + first_row, column=col)
        cell.value = d[row]
        if header:
            cell.font = Font(bold=True)
        if border:
            cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=wrap)
        cell.number_format = number_format


def write_row_to_sheet(
    sheet,
    d,
    row,
    first_col,
    header=False,
    border=False,
    wrap=False,
    number_format="#,##0.00",
):
    if isinstance(d, ndarray):
        length = d.size
    else:
        length = len(d)
    for col in range(length):
        cell = sheet.cell(row=row, column=first_col + col)
        cell.value = d[col]
        if header:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right", wrap_text=wrap)
        else:
            cell.alignment = Alignment(horizontal="center", wrap_text=wrap)
        if border:
            cell.border = thin_border
        cell.number_format = number_format


def write_table_to_sheet(sheet, d, first_row, first_col, number_format="#,##0.00"):
    for row in range(d.shape[0]):
        for col in range(d.shape[1]):
            cell = sheet.cell(row=row + first_row, column=col + first_col)
            cell.value = d[row, col]
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = number_format
