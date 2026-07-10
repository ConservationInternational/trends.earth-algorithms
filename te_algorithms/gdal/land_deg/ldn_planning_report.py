"""Reporting for LDN Planning Tool results.

Generates:
  - Excel workbook (ARR summary, hotspot ranking, BAU projection, scenario
    balance sheet).
  - JSON report for machine-readable output.
"""

import datetime as dt
import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from ... import __version__
from .. import xl

logger = logging.getLogger(__name__)

_bold = Font(bold=True)
_h_align = Alignment(horizontal="center", wrap_text=True)
_c_align = Alignment(horizontal="center")


def _hcell(sheet, row, col, val):
    c = sheet.cell(row=row, column=col, value=val)
    c.font = _bold
    c.alignment = _h_align
    c.border = xl.thin_border


def _dcell(sheet, row, col, val, fmt=None):
    c = sheet.cell(row=row, column=col, value=val)
    c.alignment = _c_align
    c.border = xl.thin_border
    if fmt:
        c.number_format = fmt


# ---------------------------------------------------------------------------
# ARR Summary sheet
# ---------------------------------------------------------------------------


def _write_arr_sheet(sheet, arr_summary: Dict[str, Any]) -> None:
    sheet.title = "ARR Summary"
    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet)
    sheet.cell(
        row=1, column=1, value="LDN Planning — Avoid/Reduce/Reverse Summary"
    ).font = Font(bold=True, size=14)
    headers = [
        "ARR Class",
        "Area (km²)",
        "% of Total",
    ]
    for col, h in enumerate(headers, 1):
        _hcell(sheet, 4, col, h)

    total = arr_summary.get("total_km2", 1) or 1
    rows = [
        ("Avoid", arr_summary.get("avoid_km2", 0)),
        ("Reduce", arr_summary.get("reduce_km2", 0)),
        ("Reverse", arr_summary.get("reverse_km2", 0)),
    ]
    for r, (label, area) in enumerate(rows, start=5):
        _dcell(sheet, r, 1, label)
        _dcell(sheet, r, 2, round(area, 2), "#,##0.00")
        _dcell(sheet, r, 3, round(100 * area / total, 2), "0.00")

    note_row = 5 + len(rows) + 1
    note = (
        "Note: Only Reverse (restoration) actions generate counterbalancing gains. "
        "Avoid and Reduce prevent losses but do not generate gains "
        "(Cowie et al. 2018; GPG Addendum 2025)."
    )
    sheet.cell(row=note_row, column=1, value=note).alignment = Alignment(wrap_text=True)
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 14


# ---------------------------------------------------------------------------
# Hotspot sheet
# ---------------------------------------------------------------------------


def _write_hotspot_sheet(
    sheet, hotspot_zones: Optional[List[Dict[str, Any]]] = None
) -> None:
    sheet.title = "Hotspot Ranking"
    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet)
    sheet.cell(row=1, column=1, value="Degradation Hotspot Ranking").font = Font(
        bold=True, size=14
    )
    headers = [
        "Priority Rank",
        "Zone ID",
        "Total Pixels",
        "Degraded Pixels",
        "Degraded Area (km²)",
        "Degraded Fraction (%)",
    ]
    for col, h in enumerate(headers, 1):
        _hcell(sheet, 4, col, h)

    if not hotspot_zones:
        sheet.cell(row=5, column=1, value="No hotspot data available.")
        return

    for row, zone in enumerate(hotspot_zones, start=5):
        _dcell(sheet, row, 1, zone.get("priority_rank"))
        _dcell(sheet, row, 2, zone.get("zone_id", zone.get("fid", "")))
        _dcell(sheet, row, 3, zone.get("total_pixels"))
        _dcell(sheet, row, 4, zone.get("deg_pixels"))
        _dcell(sheet, row, 5, round(zone.get("deg_area_km2", 0), 2), "#,##0.00")
        _dcell(
            sheet,
            row,
            6,
            round(100 * zone.get("deg_fraction", 0), 2),
            "0.00",
        )


# ---------------------------------------------------------------------------
# BAU sheet
# ---------------------------------------------------------------------------


def _write_bau_sheet(sheet, bau_summary: Dict[str, Any]) -> None:
    sheet.title = "BAU Projection"
    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet)
    sheet.cell(
        row=1, column=1, value="Business-As-Usual Degradation Projection"
    ).font = Font(bold=True, size=14)

    target_year = bau_summary.get("target_year", "")
    rows = [
        ("Year (start of baseline)", bau_summary.get("year_initial")),
        ("Year (end of baseline / reporting)", bau_summary.get("year_final")),
        ("Projection target year", target_year),
        ("Total land area (km²)", bau_summary.get("total_area_km2")),
        (
            "Degraded area — baseline (km²)",
            bau_summary.get("degraded_area_baseline_km2"),
        ),
        ("Degraded fraction — baseline (%)", bau_summary.get("pct_degraded_baseline")),
        (
            "Degraded area — reporting (km²)",
            bau_summary.get("degraded_area_reporting_km2"),
        ),
        (
            "Degraded fraction — reporting (%)",
            bau_summary.get("pct_degraded_reporting"),
        ),
        (
            "Annual net change in degraded area (km²/yr)",
            bau_summary.get("annual_change_km2"),
        ),
        (
            f"BAU projection {target_year} (km²)",
            bau_summary.get(f"bau_projection_{target_year}_km2"),
        ),
        (
            "LDN target (= baseline degraded area, km²)",
            bau_summary.get("ldntarget_km2"),
        ),
        ("Shortfall above baseline target (km²)", bau_summary.get("shortfall_km2")),
    ]

    _hcell(sheet, 3, 1, "Metric")
    _hcell(sheet, 3, 2, "Value")
    for r, (label, val) in enumerate(rows, start=4):
        sheet.cell(row=r, column=1, value=label).alignment = Alignment(wrap_text=True)
        if isinstance(val, float):
            _dcell(sheet, r, 2, round(val, 2), "#,##0.00")
        else:
            _dcell(sheet, r, 2, val)

    note_row = 4 + len(rows) + 1
    sheet.cell(
        row=note_row,
        column=1,
        value=(
            "LDN frame of reference: the target equals the baseline (no net loss). "
            "Neutrality is the minimum objective; more ambitious targets (net gain) are encouraged."
        ),
    ).alignment = Alignment(wrap_text=True)
    sheet.column_dimensions["A"].width = 45
    sheet.column_dimensions["B"].width = 20


def _write_bau_zones_sheet(
    sheet, zones: List[Dict[str, Any]], target_year: int
) -> None:
    """Per-zone BAU statistics sheet."""
    sheet.title = "BAU by Zone"
    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet)
    sheet.cell(
        row=1, column=1, value="BAU Projection — Per-Zone Statistics"
    ).font = Font(bold=True, size=14)
    headers = [
        "Zone",
        "Total Area (km²)",
        "Degraded — Baseline (km²)",
        "Degraded — Baseline (%)",
        "Degraded — Reporting (km²)",
        "Degraded — Reporting (%)",
        "Annual Change (km²/yr)",
        f"BAU Projection {target_year} (km²)",
        "LDN Target (km²)",
        "Shortfall (km²)",
    ]
    for col, h in enumerate(headers, 1):
        _hcell(sheet, 4, col, h)

    if not zones:
        sheet.cell(row=5, column=1, value="No zone data available.")
        return

    proj_key = f"bau_projection_{target_year}_km2"
    for row, z in enumerate(zones, start=5):
        _dcell(sheet, row, 1, z.get("zone_name", ""))
        _dcell(sheet, row, 2, round(z.get("total_area_km2", 0) or 0, 2), "#,##0.00")
        _dcell(
            sheet,
            row,
            3,
            round(z.get("degraded_area_baseline_km2", 0) or 0, 2),
            "#,##0.00",
        )
        _dcell(sheet, row, 4, round(z.get("pct_degraded_baseline", 0) or 0, 2), "0.00")
        v = z.get("degraded_area_reporting_km2")
        _dcell(sheet, row, 5, round(v, 2) if v is not None else "", "#,##0.00")
        v = z.get("pct_degraded_reporting")
        _dcell(sheet, row, 6, round(v, 2) if v is not None else "", "0.00")
        v = z.get("annual_change_km2")
        _dcell(sheet, row, 7, round(v, 3) if v is not None else "", "#,##0.000")
        v = z.get(proj_key)
        _dcell(sheet, row, 8, round(v, 2) if v is not None else "", "#,##0.00")
        _dcell(sheet, row, 9, round(z.get("ldntarget_km2", 0) or 0, 2), "#,##0.00")
        v = z.get("shortfall_km2")
        _dcell(sheet, row, 10, round(v, 2) if v is not None else "", "#,##0.00")

    for col_letter, width in zip(
        "ABCDEFGHIJ", [28, 14, 22, 20, 22, 20, 20, 22, 14, 14]
    ):
        sheet.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# Scenario sheet
# ---------------------------------------------------------------------------


def _write_scenario_sheet(sheet, scenario_summary: Dict[str, Any]) -> None:
    sheet.title = "Scenario Balance Sheet"
    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet)
    sheet.cell(
        row=1, column=1, value="LDN Planning — Scenario Balance Sheet"
    ).font = Font(bold=True, size=14)

    # Per-target table
    per_target = scenario_summary.get("per_target", [])
    if per_target:
        headers = [
            "Target #",
            "Intervention",
            "Effectiveness",
            "Area Treated (km²)",
            "Gains (km²)",
            "Avoided Losses (km²)",
        ]
        for col, h in enumerate(headers, 1):
            _hcell(sheet, 4, col, h)
        for r, t in enumerate(per_target, start=5):
            _dcell(sheet, r, 1, t.get("target_index", r - 5) + 1)
            _dcell(sheet, r, 2, t.get("intervention", "").capitalize())
            _dcell(sheet, r, 3, round(t.get("effectiveness", 0) * 100, 1), "0.0")
            _dcell(sheet, r, 4, round(t.get("area_treated_km2", 0), 2), "#,##0.00")
            _dcell(sheet, r, 5, round(t.get("gains_km2", 0), 2), "#,##0.00")
            _dcell(sheet, r, 6, round(t.get("avoided_losses_km2", 0), 2), "#,##0.00")

        sum_row = 5 + len(per_target) + 1
    else:
        sum_row = 5

    # Totals
    total_start = sum_row
    totals = [
        (
            "Gains from Reverse (counterbalancing) (km²)",
            scenario_summary.get("gains_km2_reverse", 0),
        ),
        (
            "Avoided losses from Reduce (km²)",
            scenario_summary.get("avoided_losses_km2_reduce", 0),
        ),
        (
            "Avoided losses from Avoid (km²)",
            scenario_summary.get("avoided_losses_km2_avoid", 0),
        ),
        ("Total gains (km²)", scenario_summary.get("total_gains_km2", 0)),
        (
            "Total avoided losses (km²)",
            scenario_summary.get("total_avoided_losses_km2", 0),
        ),
    ]
    _hcell(sheet, total_start, 1, "Summary")
    _hcell(sheet, total_start, 2, "km²")
    for r, (label, val) in enumerate(totals, start=total_start + 1):
        sheet.cell(row=r, column=1, value=label)
        _dcell(sheet, r, 2, round(val, 2), "#,##0.00")

    note_row = total_start + len(totals) + 2
    sheet.cell(
        row=note_row, column=1, value=scenario_summary.get("net_balance_note", "")
    ).alignment = Alignment(wrap_text=True)
    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 16


def _write_scenario_breakdown_sheet(
    sheet, title: str, rows: List[Dict[str, Any]], first_col_header: str
) -> None:
    """Per-land-type scenario breakdown sheet."""
    sheet.title = title
    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet)
    sheet.cell(row=1, column=1, value=f"LDN Planning — {title}").font = Font(
        bold=True, size=14
    )
    headers = [
        first_col_header,
        "Gains — Reverse (km²)",
        "Avoided Losses — Reduce (km²)",
        "Avoided Losses — Avoid (km²)",
        "Total Gains (km²)",
        "Total Avoided Losses (km²)",
    ]
    for col, h in enumerate(headers, 1):
        _hcell(sheet, 4, col, h)

    if not rows:
        sheet.cell(row=5, column=1, value="No data available.")
        return

    for r, item in enumerate(rows, start=5):
        _dcell(sheet, r, 1, item.get("name", ""))
        _dcell(sheet, r, 2, round(item.get("gains_km2", 0) or 0, 2), "#,##0.00")
        _dcell(
            sheet,
            r,
            3,
            round(item.get("avoided_losses_reduce_km2", 0) or 0, 2),
            "#,##0.00",
        )
        _dcell(
            sheet,
            r,
            4,
            round(item.get("avoided_losses_avoid_km2", 0) or 0, 2),
            "#,##0.00",
        )
        _dcell(sheet, r, 5, round(item.get("total_gains_km2", 0) or 0, 2), "#,##0.00")
        _dcell(
            sheet,
            r,
            6,
            round(item.get("total_avoided_losses_km2", 0) or 0, 2),
            "#,##0.00",
        )

    note_row = 5 + len(rows) + 1
    sheet.cell(
        row=note_row,
        column=1,
        value=(
            "Only Reverse gains count toward LDN counterbalancing; Avoid/Reduce "
            "are avoided losses (upper bound). Per-land-type accounting mirrors "
            "the GPG Addendum 'like for like' principle."
        ),
    ).alignment = Alignment(wrap_text=True)
    for col_letter, width in zip("ABCDEF", [30, 20, 26, 26, 18, 24]):
        sheet.column_dimensions[col_letter].width = width


def _write_projection_summary_sheet(sheet, projection: Dict[str, Any]) -> None:
    """BAU-vs-scenario comparison over the planning horizon."""
    sheet.title = "BAU vs Scenario"
    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet)
    sheet.cell(row=1, column=1, value="LDN Planning — BAU vs Scenario").font = Font(
        bold=True, size=14
    )

    target_year = projection.get("target_year", "")
    rows = [
        ("Baseline degraded area (km²)", projection.get("degraded_area_baseline_km2")),
        (
            "Reporting degraded area (km²)",
            projection.get("degraded_area_reporting_km2"),
        ),
        (f"BAU projection {target_year} (km²)", projection.get("bau_projection_km2")),
        ("LDN target (= baseline, km²)", projection.get("ldntarget_km2")),
        ("Scenario gains — Reverse (km²)", projection.get("scenario_gains_km2")),
        (
            "Scenario avoided losses — Avoid+Reduce (km²)",
            projection.get("scenario_avoided_losses_km2"),
        ),
        (
            "Total scenario contribution (km²)",
            projection.get("scenario_contribution_km2"),
        ),
        (
            f"Scenario-adjusted degraded area {target_year} (km²)",
            projection.get("scenario_degraded_km2"),
        ),
        ("BAU shortfall above target (km²)", projection.get("bau_shortfall_km2")),
        (
            "Remaining shortfall after scenario (km²)",
            projection.get("remaining_shortfall_km2"),
        ),
        ("Shortfall gap closed (%)", projection.get("gap_closed_pct")),
        ("Neutrality achieved at target year", projection.get("neutral")),
    ]
    _hcell(sheet, 3, 1, "Metric")
    _hcell(sheet, 3, 2, "Value")
    for r, (label, val) in enumerate(rows, start=4):
        sheet.cell(row=r, column=1, value=label).alignment = Alignment(wrap_text=True)
        if isinstance(val, bool):
            _dcell(sheet, r, 2, "Yes" if val else "No")
        elif isinstance(val, float):
            _dcell(sheet, r, 2, round(val, 2), "#,##0.00")
        else:
            _dcell(sheet, r, 2, val)

    note_row = 4 + len(rows) + 1
    sheet.cell(
        row=note_row, column=1, value=projection.get("note", "")
    ).alignment = Alignment(wrap_text=True)
    sheet.column_dimensions["A"].width = 48
    sheet.column_dimensions["B"].width = 20

    # Optional per-zone neutrality table
    by_zone = projection.get("by_zone") or []
    if by_zone:
        start = note_row + 2
        headers = [
            "Land Type",
            f"BAU Projection {target_year} (km²)",
            "LDN Target (km²)",
            "Scenario Contribution (km²)",
            f"Scenario-Adjusted {target_year} (km²)",
            "Gap Closed (%)",
            "Neutral",
        ]
        for col, h in enumerate(headers, 1):
            _hcell(sheet, start, col, h)
        for r, z in enumerate(by_zone, start=start + 1):
            _dcell(sheet, r, 1, z.get("zone_name", ""))
            _dcell(sheet, r, 2, round(z.get("bau_projection_km2", 0), 2), "#,##0.00")
            _dcell(sheet, r, 3, round(z.get("ldntarget_km2", 0), 2), "#,##0.00")
            _dcell(
                sheet,
                r,
                4,
                round(z.get("scenario_contribution_km2", 0), 2),
                "#,##0.00",
            )
            _dcell(sheet, r, 5, round(z.get("scenario_degraded_km2", 0), 2), "#,##0.00")
            _dcell(sheet, r, 6, round(z.get("gap_closed_pct", 0), 1), "0.0")
            _dcell(sheet, r, 7, "Yes" if z.get("neutral") else "No")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def save_ldn_planning_excel(
    output_path: Path,
    arr_summary: Optional[Dict[str, Any]] = None,
    hotspot_zones: Optional[List[Dict[str, Any]]] = None,
    bau_summary: Optional[Dict[str, Any]] = None,
    scenario_summary: Optional[Dict[str, Any]] = None,
    projection_summary: Optional[Dict[str, Any]] = None,
) -> None:
    """Save LDN Planning results to an Excel workbook.

    Only sheets for which data are provided are written.
    """
    wb = Workbook()
    first = True

    def _next_sheet(title: str):
        nonlocal first
        if first:
            ws = wb.active
            ws.title = title
            first = False
            return ws
        return wb.create_sheet(title)

    if arr_summary is not None:
        _write_arr_sheet(_next_sheet("ARR Summary"), arr_summary)
    if hotspot_zones is not None:
        _write_hotspot_sheet(_next_sheet("Hotspot Ranking"), hotspot_zones)
    if bau_summary is not None:
        _write_bau_sheet(_next_sheet("BAU Projection"), bau_summary)
        bau_zones = bau_summary.get("zones")
        if bau_zones:
            _write_bau_zones_sheet(
                _next_sheet("BAU by Zone"),
                bau_zones,
                bau_summary.get("target_year", ""),
            )
    if scenario_summary is not None:
        _write_scenario_sheet(_next_sheet("Scenario Balance Sheet"), scenario_summary)
        by_land_type = scenario_summary.get("by_land_type")
        if by_land_type:
            _write_scenario_breakdown_sheet(
                _next_sheet("Scenario by Land Type"),
                "Scenario by Land Type",
                by_land_type,
                "Land Type",
            )
        by_zone = scenario_summary.get("by_zone")
        if by_zone:
            _write_scenario_breakdown_sheet(
                _next_sheet("Scenario by Land Type"),
                "Scenario by Land Type",
                by_zone,
                "Land Type",
            )
    if projection_summary is not None:
        _write_projection_summary_sheet(
            _next_sheet("BAU vs Scenario"), projection_summary
        )

    try:
        wb.save(str(output_path))
        logger.info("LDN Planning table saved to %s", output_path)
    except OSError:
        logger.error(
            "Error saving LDN Planning table — check %s is not already open.",
            output_path,
        )


def save_ldn_planning_json(
    output_path: Path,
    task_name: str = "LDN Planning",
    arr_summary: Optional[Dict[str, Any]] = None,
    hotspot_summary: Optional[Dict[str, Any]] = None,
    bau_summary: Optional[Dict[str, Any]] = None,
    scenario_summary: Optional[Dict[str, Any]] = None,
    projection_summary: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Save LDN Planning results to a JSON file and return the report dict."""
    report = {
        "task_name": task_name,
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "te_algorithms_version": __version__,
        "arr": arr_summary,
        "hotspots": hotspot_summary,
        "bau": bau_summary,
        "scenario": scenario_summary,
        "projection": projection_summary,
    }
    output_path = Path(output_path)
    try:
        output_path.write_text(
            json.dumps(report, default=str, indent=2), encoding="utf-8"
        )
        logger.info("LDN Planning JSON saved to %s", output_path)
    except OSError:
        logger.error("Error saving LDN Planning JSON to %s", output_path)
    return report
