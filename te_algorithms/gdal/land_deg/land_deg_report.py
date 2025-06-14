import datetime as dt
import json
import logging
from pathlib import Path
from typing import TYPE_CHECKING, Dict, List, Union

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils.cell import get_column_letter
from te_schemas import land_cover, reporting, schemas

from ... import __release_date__, __version__
from .. import xl
from . import config, models

if TYPE_CHECKING:
    from te_schemas.aoi import AOI

logger = logging.getLogger(__name__)


def save_summary_table_excel(
    output_path: Path,
    summary_table: models.SummaryTableLD,
    periods,
    land_cover_years: List[int],
    soil_organic_carbon_years: List[int],
    lc_legend_nesting: land_cover.LCLegendNesting,
    lc_trans_matrix: land_cover.LCTransitionDefinitionDeg,
    period_name,
):
    """Save summary table into an xlsx file on disk"""
    template_summary_table_path = (
        Path(__file__).parents[2] / "data/summary_table_ld_sdg.xlsx"
    )
    workbook = load_workbook(str(template_summary_table_path))
    _render_ld_workbook(
        workbook,
        summary_table,
        periods,
        land_cover_years,
        soil_organic_carbon_years,
        lc_legend_nesting,
        lc_trans_matrix,
        period_name,
    )
    try:
        workbook.save(output_path)
        logger.info("Indicator table saved to {}".format(output_path))

    except OSError:
        error_message = (
            f"Error saving output table - check that {output_path!r} is accessible "
            f"and not already open."
        )
        logger.error(error_message)


def _get_population_list_by_degradation_class(pop_by_deg_class, pop_type):
    return reporting.PopulationList(
        "Population by degradation class",
        [
            reporting.Population("Improved", pop_by_deg_class.get(1, 0), type=pop_type),
            reporting.Population("Stable", pop_by_deg_class.get(0, 0), type=pop_type),
            reporting.Population(
                "Degraded", pop_by_deg_class.get(-1, 0), type=pop_type
            ),
            reporting.Population(
                "No data", pop_by_deg_class.get(config.NODATA_VALUE, 0), type=pop_type
            ),
        ],
    )


def save_reporting_json(
    output_path: Path,
    summary_tables: Dict[str, models.SummaryTableLD],
    summary_table_status: Union[None, models.SummaryTableLD],
    summary_table_change: Union[None, models.SummaryTableLDChange],
    params: dict,
    task_name: str,
    aoi: "AOI",
    summary_table_kwargs: dict,
):
    land_condition_reports = {}
    affected_pop_reports = {}

    assert len(summary_tables) == len(params["periods"])

    map_legend = {
        "Improved": 1,
        "Stable": 0,
        "Degraded": -1,
        "No data": config.NODATA_VALUE,
    }

    for period_num, period in enumerate(params["periods"]):
        period_name = period["name"]
        period_params = period["params"]
        st = summary_tables[period_name]

        ##########################################################################
        # Area summary tables
        lc_legend_nesting = summary_table_kwargs[period_name]["lc_legend_nesting"]
        lc_trans_matrix = summary_table_kwargs[period_name]["lc_trans_matrix"]

        period_sdg_summary = reporting.AreaList(
            "SDG Indicator 15.3.1",
            "sq km",
            [
                reporting.Area(class_name, st.sdg_summary.get(code, 0.0))
                for class_name, code in map_legend.items()
            ],
        )

        prod_summaries = {
            key: reporting.AreaList(
                "Productivity",
                "sq km",
                [
                    reporting.Area(class_name, value.get(code, 0.0))
                    for class_name, code in map_legend.items()
                ],
            )
            for key, value in st.prod_summary.items()
        }

        lc_summary = reporting.AreaList(
            "Land cover",
            "sq km",
            [
                reporting.Area(class_name, st.lc_summary.get(code, 0.0))
                for class_name, code in map_legend.items()
            ],
        )

        soc_summaries = {
            key: reporting.AreaList(
                "Soil organic carbon",
                "sq km",
                [
                    reporting.Area(class_name, value.get(code, 0.0))
                    for class_name, code in map_legend.items()
                ],
            )
            for key, value in st.soc_summary.items()
        }

        #######################################################################
        # Productivity tables
        lc_trans_dict = lc_trans_matrix.get_transition_integers_key()

        crosstab_prod = []
        # If no land cover data was available for first year of productivity
        # data, then won't be able to output these tables, so need to check
        # first if data is available

        if len([*st.lc_trans_prod_bizonal.keys()]) > 0:
            for prod_name, prod_code in config.PRODUCTIVITY_CLASS_KEY.items():
                crosstab_entries = []

                for transition, codes in lc_trans_dict.items():
                    initial_class = lc_trans_matrix.legend.classByCode(
                        codes["initial"]
                    ).get_name()
                    final_class = lc_trans_matrix.legend.classByCode(
                        codes["final"]
                    ).get_name()
                    crosstab_entries.append(
                        reporting.CrossTabEntry(
                            initial_class,
                            final_class,
                            value=st.lc_trans_prod_bizonal.get(
                                (transition, prod_code), 0.0
                            ),
                        )
                    )
                crosstab_prod.append(
                    reporting.CrossTab(
                        prod_name,
                        unit="sq km",
                        initial_year=period_params["periods"]["productivity"][
                            "year_initial"
                        ],
                        final_year=period_params["periods"]["productivity"][
                            "year_final"
                        ],
                        values=crosstab_entries,
                    )
                )

        #######################################################################
        # Land cover tables
        land_cover_years = period_params["layer_lc_years"]

        ###
        # LC transition cross tabs
        crosstab_lcs = []

        for lc_trans_zonal_areas, lc_trans_zonal_areas_period in zip(
            st.lc_trans_zonal_areas, st.lc_trans_zonal_areas_periods
        ):
            lc_by_transition_type = []

            for transition, codes in lc_trans_dict.items():
                initial_class = lc_trans_matrix.legend.classByCode(
                    codes["initial"]
                ).get_name()
                final_class = lc_trans_matrix.legend.classByCode(
                    codes["final"]
                ).get_name()
                lc_by_transition_type.append(
                    reporting.CrossTabEntry(
                        initial_class,
                        final_class,
                        value=lc_trans_zonal_areas.get(transition, 0.0),
                    )
                )
            lc_by_transition_type = sorted(
                lc_by_transition_type, key=lambda i: i.value, reverse=True
            )
            crosstab_lc = reporting.CrossTab(
                name="Land area by land cover transition type",
                unit="sq km",
                initial_year=lc_trans_zonal_areas_period["year_initial"],
                final_year=lc_trans_zonal_areas_period["year_final"],
                # TODO: Check indexing as may be missing a class
                values=lc_by_transition_type,
            )
            crosstab_lcs.append(crosstab_lc)

        ###
        # LC by year
        lc_by_year = {}

        for year_num, year in enumerate(land_cover_years):
            # total_land_area = sum(
            #     [
            #         value
            #         for key, value in st.lc_annual_totals[year_num].items()
            #         if key != config.MASK_VALUE
            #     ]
            # )
            # logging.debug(
            #     f"Total land area in {year} per land cover data {total_land_area}"
            # )
            #
            # for lc_class in lc_trans_matrix.legend.key:
            #     logging.debug(
            #         f"Total area of {lc_class.get_name()} in {year}: {st.lc_annual_totals[year_num].get(lc_class.code, 0.0)}"
            #     )

            lc_by_year[int(year)] = {
                lc_class.get_name(): st.lc_annual_totals[year_num].get(
                    lc_class.code, 0.0
                )
                for lc_class in lc_trans_matrix.legend.key
                + [lc_trans_matrix.legend.nodata]
            }
        lc_by_year_by_class = reporting.ValuesByYearDict(
            name="Area by year by land cover class", unit="sq km", values=lc_by_year
        )

        #######################################################################
        # Soil organic carbon tables
        soil_organic_carbon_years = period_params["layer_soc_years"]

        ###
        # SOC by year by land cover class
        soc_by_year = {}

        for year_num, soc_by_lc_annual_total in enumerate(st.soc_by_lc_annual_totals):
            year = soil_organic_carbon_years[year_num]
            soc_by_year[int(year)] = {
                lc_class.get_name(): soc_by_lc_annual_total.get(lc_class.code, 0.0)
                for lc_class in lc_trans_matrix.legend.key
                + [lc_trans_matrix.legend.nodata]
            }

        soc_by_year_by_class = reporting.ValuesByYearDict(
            name="Soil organic carbon by year by land cover class",
            unit="tonnes",
            values=soc_by_year,
        )

        ###
        # Setup this period's land condition report
        period_assessment = reporting.LandConditionAssessment(
            sdg=reporting.SDG15Report(summary=period_sdg_summary),
            productivity=reporting.ProductivityReport(
                summaries=prod_summaries,
                crosstabs_by_productivity_class=crosstab_prod,
            ),
            land_cover=reporting.LandCoverReport(
                summary=lc_summary,
                legend_nesting=lc_legend_nesting,
                transition_matrix=lc_trans_matrix,
                crosstabs_by_land_cover_class=crosstab_lcs,
                land_cover_areas_by_year=lc_by_year_by_class,
            ),
            soil_organic_carbon=reporting.SoilOrganicCarbonReport(
                summaries=soc_summaries,
                soc_stock_by_year=soc_by_year_by_class,
            ),
        )

        ###
        # Setup this period's affected population report

        affected_by_deg_summary = {
            "Total population": _get_population_list_by_degradation_class(
                st.sdg_zonal_population_total, "Total population"
            )
        }

        if len(st.sdg_zonal_population_female) > 0:
            affected_by_deg_summary["Female population"] = (
                _get_population_list_by_degradation_class(
                    st.sdg_zonal_population_female, "Female population"
                )
            )

        if len(st.sdg_zonal_population_male) > 0:
            affected_by_deg_summary["Male population"] = (
                _get_population_list_by_degradation_class(
                    st.sdg_zonal_population_male, "Male population"
                )
            )

        affected_pop_reports[period_name] = reporting.AffectedPopulationReport(
            affected_by_deg_summary
        )

        if period_num == 0:
            status_assessment = None
            change_assessment = None
        else:
            assert summary_table_status is not None
            assert summary_table_change is not None

            # There is one less status than there are periods (given status is a
            # comparison of two successive periods)
            summary_number = period_num - 1
            status_assessment = reporting.LandConditionStatus(
                sdg=reporting.AreaList(
                    "SDG Indicator 15.3.1 (progress since baseline)",
                    "sq km",
                    [
                        reporting.Area(
                            class_name,
                            summary_table_status.sdg_summaries[summary_number].get(
                                code, 0.0
                            ),
                        )
                        for class_name, code in map_legend.items()
                    ],
                ),
                productivity={
                    key: reporting.AreaList(
                        "Productivity (progress since baseline)",
                        "sq km",
                        [
                            reporting.Area(class_name, value.get(code, 0.0))
                            for class_name, code in map_legend.items()
                        ],
                    )
                    for key, value in summary_table_status.prod_summaries[
                        summary_number
                    ].items()
                },
                land_cover=reporting.AreaList(
                    "Land cover (progress since baseline)",
                    "sq km",
                    [
                        reporting.Area(
                            class_name,
                            summary_table_status.lc_summaries[summary_number].get(
                                code, 0.0
                            ),
                        )
                        for class_name, code in map_legend.items()
                    ],
                ),
                soil_organic_carbon={
                    key: reporting.AreaList(
                        "Soil organic carbon (progress since baseline)",
                        "sq km",
                        [
                            reporting.Area(class_name, value.get(code, 0.0))
                            for class_name, code in map_legend.items()
                        ],
                    )
                    for key, value in summary_table_status.soc_summaries[
                        summary_number
                    ].items()
                },
            )

            ### Change report

            change_assessment = reporting.LandConditionChange(
                sdg=reporting.CrossTab(
                    name="Land area by change in degradation status (one-out all-out layer)",
                    unit="sq km",
                    initial_year=period_params["periods"]["productivity"][
                        "year_initial"
                    ],
                    final_year=period_params["periods"]["productivity"]["year_final"],
                    values=[
                        reporting.CrossTabEntry(
                            initial_class_name,
                            final_class_name,
                            value=summary_table_change.sdg_crosstabs[
                                summary_number
                            ].get((initial_code, final_code), 0.0),
                        )
                        for initial_class_name, initial_code in map_legend.items()
                        for final_class_name, final_code in map_legend.items()
                    ],
                ),
                productivity=reporting.CrossTab(
                    name="Land area by change in land productivity degradation status",
                    unit="sq km",
                    initial_year=period_params["periods"]["productivity"][
                        "year_initial"
                    ],
                    final_year=period_params["periods"]["productivity"]["year_final"],
                    values=[
                        reporting.CrossTabEntry(
                            initial_class_name,
                            final_class_name,
                            value=summary_table_change.prod_crosstabs[
                                summary_number
                            ].get((initial_code, final_code), 0.0),
                        )
                        for initial_class_name, initial_code in map_legend.items()
                        for final_class_name, final_code in map_legend.items()
                    ],
                ),
                land_cover=reporting.CrossTab(
                    name="Land area by change in land cover degradation status",
                    unit="sq km",
                    initial_year=period_params["periods"]["land_cover"]["year_initial"],
                    final_year=period_params["periods"]["land_cover"]["year_final"],
                    values=[
                        reporting.CrossTabEntry(
                            initial_class_name,
                            final_class_name,
                            value=summary_table_change.lc_crosstabs[summary_number].get(
                                (initial_code, final_code), 0.0
                            ),
                        )
                        for initial_class_name, initial_code in map_legend.items()
                        for final_class_name, final_code in map_legend.items()
                    ],
                ),
                soil_organic_carbon=reporting.CrossTab(
                    name="Land area by change in soil organic carbon degradation status",
                    unit="sq km",
                    initial_year=period_params["periods"]["soc"]["year_initial"],
                    final_year=period_params["periods"]["soc"]["year_final"],
                    values=[
                        reporting.CrossTabEntry(
                            initial_class_name,
                            final_class_name,
                            value=summary_table_change.soc_crosstabs[
                                summary_number
                            ].get((initial_code, final_code), 0.0),
                        )
                        for initial_class_name, initial_code in map_legend.items()
                        for final_class_name, final_code in map_legend.items()
                    ],
                ),
            )

        land_condition_reports[period_name] = reporting.LandConditionReport(
            period_assessment=period_assessment,
            status_assessment=status_assessment,
            change_assessment=change_assessment,
        )

    ##########################################################################
    # Format final JSON output
    te_summary = reporting.TrendsEarthLandConditionSummary(
        metadata=reporting.ReportMetadata(
            title="Trends.Earth Summary Report",
            date=dt.datetime.now(dt.timezone.utc),
            trends_earth_version=schemas.TrendsEarthVersion(
                version=__version__,
                revision=None,
                release_date=dt.datetime.strptime(
                    __release_date__, "%Y/%m/%d %H:%M:%SZ"
                ),
            ),
            area_of_interest=schemas.AreaOfInterest(
                name=task_name,  # TODO replace this with area of interest name once implemented in TE
                geojson=aoi.get_geojson(),
                crs_wkt=aoi.get_crs_wkt(),
            ),
        ),
        land_condition=land_condition_reports,
        affected_population=affected_pop_reports,
    )

    try:
        te_summary_json = json.loads(
            reporting.TrendsEarthLandConditionSummary.Schema().dumps(te_summary)
        )
        with open(output_path, "w") as f:
            json.dump(te_summary_json, f, indent=4)

        return te_summary_json

    except OSError:
        logger.error(
            "Error saving indicator table JSON - check that "
            f"{output_path} is accessible and not already open."
        )

        return None


def _render_ld_workbook(
    template_workbook,
    summary_table: models.SummaryTableLD,
    periods,
    lc_years: List[int],
    soc_years: List[int],
    lc_legend_nesting: land_cover.LCLegendNesting,
    lc_trans_matrix: land_cover.LCTransitionDefinitionDeg,
    period_name,
):
    _write_overview_sheet(template_workbook["SDG 15.3.1"], summary_table)
    _write_productivity_sheet(
        template_workbook["Productivity"], summary_table, lc_trans_matrix
    )
    _write_soc_sheet(
        template_workbook["Soil organic carbon"],
        summary_table,
        lc_trans_matrix,
        lc_legend_nesting,
    )
    _write_land_cover_sheet(
        template_workbook["Land cover"],
        summary_table,
        lc_trans_matrix,
        periods["land_cover"],
    )
    _write_population_sheet(template_workbook["Population"], summary_table)

    return template_workbook


def _get_summary_array(d):
    """pulls summary values for excel sheet from a summary dictionary"""

    return np.array([d.get(1, 0), d.get(0, 0), d.get(-1, 0), d.get(-32768, 0)])


def _write_overview_sheet(sheet, summary_table: models.SummaryTableLD):
    xl.write_col_to_sheet(sheet, _get_summary_array(summary_table.sdg_summary), 6, 6)
    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet)


def _write_productivity_sheet(
    sheet,
    st: models.SummaryTableLD,
    lc_trans_matrix: land_cover.LCTransitionDefinitionDeg,
):
    xl.write_col_to_sheet(
        sheet, _get_summary_array(st.prod_summary["all_cover_types"]), 6, 6
    )
    classes = [c.get_name() for c in lc_trans_matrix.legend.key]

    if len(classes) > 7:
        # Land cover tables by default only have room for 7 classes. So need to
        # add more columns/rows if there are more than 7 land cover classes
        sheet.insert_cols(8, len(classes) - 7)

    if len([*st.lc_trans_prod_bizonal.keys()]) > 0:
        # If no land cover data was available for first year of productivity
        # data, then won't be able to output these tables
        if len(classes) > 7:
            n_new_rows_per_table = len(classes) - 7
        else:
            n_new_rows_per_table = 0
        write_crosstab_by_lc(
            sheet,
            _get_prod_table(st.lc_trans_prod_bizonal, 5, lc_trans_matrix),
            classes,
            13,
            1,
        )
        write_crosstab_by_lc(
            sheet,
            _get_prod_table(st.lc_trans_prod_bizonal, 4, lc_trans_matrix),
            classes,
            25 + n_new_rows_per_table,
            1,
        )
        write_crosstab_by_lc(
            sheet,
            _get_prod_table(st.lc_trans_prod_bizonal, 3, lc_trans_matrix),
            classes,
            37 + n_new_rows_per_table * 2,
            1,
        )
        write_crosstab_by_lc(
            sheet,
            _get_prod_table(st.lc_trans_prod_bizonal, 2, lc_trans_matrix),
            classes,
            49 + n_new_rows_per_table * 3,
            1,
        )
        write_crosstab_by_lc(
            sheet,
            _get_prod_table(st.lc_trans_prod_bizonal, 1, lc_trans_matrix),
            classes,
            61 + n_new_rows_per_table * 4,
            1,
        )
        write_crosstab_by_lc(
            sheet,
            _get_prod_table(
                st.lc_trans_prod_bizonal, config.NODATA_VALUE, lc_trans_matrix
            ),
            classes,
            73 + n_new_rows_per_table * 5,
            1,
        )
    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet)


def _add_header_cell(sheet, row, col, value):
    cell = sheet.cell(row=row, column=col)
    cell.value = value
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = xl.thin_border
    cell.font = Font(bold=True)


def _write_soc_sheet(
    sheet,
    st: models.SummaryTableLD,
    lc_trans_matrix: land_cover.LCTransitionDefinitionDeg,
    lc_legend_nesting: land_cover.LCLegendNesting,
):
    # Exclude any codes that are nested under IPCC class 7 (water)
    excluded_codes = lc_legend_nesting.nesting[7]

    logger.info("Excluding land cover codes nested under water (%s)", excluded_codes)

    classes = [
        c.get_name() for c in lc_trans_matrix.legend.key if c.code not in excluded_codes
    ]

    if len(classes) > 6:
        # SOC stock change table by default only has room for 6 classes. So need to
        # add more columns/rows if there are more than 6 land cover classes
        sheet.insert_rows(16 + 1, len(classes) - 6)
        n_new_rows_per_table = len(classes) - 6
    else:
        n_new_rows_per_table = 0

    xl.write_col_to_sheet(sheet, _get_summary_array(st.soc_summary["non_water"]), 6, 6)

    if len(classes) > 6:
        # SOC stock change by land cover table by default only has room for 6
        # classes. So need to add more columns if there are more than 6 land
        # cover classes. More rows will be added in the setup_crosstab_by_lc
        # function. Add the extra columns now so that the tables at the top of
        # the sheet don't get messed up.
        sheet.insert_cols(8, len(classes) - 6)

    first_data_row = 16
    last_data_row = first_data_row + len(classes) - 1
    if st.soc_by_lc_annual_totals != []:
        xl.write_col_to_sheet(
            sheet,
            _get_totals_by_lc_class_as_array(
                st.soc_by_lc_annual_totals[0],
                lc_trans_matrix,
                excluded_codes=excluded_codes,
            ),
            7,
            first_data_row,
            border=True,
            number_format="#,##0",
        )
        # Now write target
        xl.write_col_to_sheet(
            sheet,
            _get_totals_by_lc_class_as_array(
                st.soc_by_lc_annual_totals[-1],
                lc_trans_matrix,
                excluded_codes=excluded_codes,
            ),
            8,
            first_data_row,
            border=True,
            number_format="#,##0",
        )

    xl.write_col_to_sheet(
        sheet, classes, 2, first_data_row, header=True, border=True, wrap=True
    )

    if st.lc_annual_totals != []:
        # Write table of baseline areas
        lc_bl_no_water = _get_totals_by_lc_class_as_array(
            st.lc_annual_totals[0], lc_trans_matrix, excluded_codes=excluded_codes
        )
        xl.write_col_to_sheet(
            sheet, lc_bl_no_water, 5, 16, border=True, number_format="#,##0"
        )
        # Write table of final year areas
        lc_final_no_water = _get_totals_by_lc_class_as_array(
            st.lc_annual_totals[-1], lc_trans_matrix, excluded_codes=excluded_codes
        )
        xl.write_col_to_sheet(
            sheet, lc_final_no_water, 6, 16, border=True, number_format="#,##0"
        )

    # Add initial/final SOC in tonnes/ha
    for row in sheet.iter_rows(
        min_row=first_data_row, max_row=last_data_row, min_col=3, max_col=4
    ):
        for cell in row:
            cell.value = (
                f"={cell.offset(column=4).column_letter}{cell.row}/"
                + f"({cell.offset(column=2).column_letter}{cell.row}*100)"
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = xl.thin_border
            cell.number_format = "#,##0.00"

    # Merge table header for SOC change table
    sheet.merge_cells(
        start_row=first_data_row - 3,
        start_column=1,
        end_row=first_data_row - 3,
        end_column=11,
    )

    # Add SOC change in tonnes
    _add_header_cell(sheet, first_data_row - 1, 8, "Final soil organic carbon (tonnes)")
    _add_header_cell(
        sheet, first_data_row - 1, 9, "Change in soil organic carbon (tonnes)"
    )
    for row in sheet.iter_rows(
        min_row=first_data_row, max_row=last_data_row, min_col=9, max_col=9
    ):
        for cell in row:
            cell.value = (
                f"={cell.offset(column=-1).column_letter}{cell.row} -"
                + f"{cell.offset(column=-2).column_letter}{cell.row}"
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = xl.thin_border
            cell.number_format = "#,##0.00"

    for row in sheet.iter_rows(
        min_row=last_data_row + 1, max_row=last_data_row + 1, min_col=5, max_col=9
    ):
        for cell in row:
            cell.value = (
                f"=sum({cell.column_letter}{first_data_row}"
                + f":{cell.column_letter}{last_data_row})"
            )
            cell.font = Font(italic=True)
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = "#,##0.00"

    # Add SOC change in percent
    _add_header_cell(
        sheet, first_data_row - 1, 10, "Change in soil organic carbon (percent)"
    )
    for row in sheet.iter_rows(
        min_row=first_data_row, max_row=last_data_row, min_col=10, max_col=10
    ):
        for cell in row:
            cell.value = (
                f"={cell.offset(column=-1).column_letter}{cell.row} /"
                + f"{cell.offset(column=-3).column_letter}{cell.row}"
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = xl.thin_border
            cell.number_format = numbers.FORMAT_PERCENTAGE

    # Set value for cell showing percent change in SOC
    sheet["G11"].value = (
        f"=(H{22 + n_new_rows_per_table} - G{22 + n_new_rows_per_table})"
        f"/G{22 + n_new_rows_per_table}"
    )

    # Merge note row at end of table
    note_row = 34 + n_new_rows_per_table * 2
    sheet.merge_cells(
        start_row=note_row, start_column=1, end_row=note_row, end_column=10
    )

    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet)


def write_crosstab_by_lc(
    sheet,
    array,
    classes,
    ul_row,
    ul_col,
):
    setup_crosstab_by_lc(sheet, classes, ul_row, ul_col)

    xl.write_table_to_sheet(sheet, array, ul_row + 3, ul_col + 2)


def setup_crosstab_by_lc(sheet, classes, ul_row, ul_col, n_classes_in_template=7):
    if len(classes) > n_classes_in_template:
        # Tables in the excel sheet by default only have room for a certain
        # number of classes. So need to add more columns/rows if there are more
        # than n_classes_in_template land cover classes. Insert
        # them starting from the second to last row of the table.
        sheet.insert_rows(
            ul_row + n_classes_in_template, len(classes) - n_classes_in_template
        )

    # Write headers
    xl.write_col_to_sheet(
        sheet, classes, ul_col + 1, ul_row + 3, header=True, wrap=True
    )
    xl.write_row_to_sheet(
        sheet, classes, ul_row + 2, ul_col + 2, header=True, wrap=True
    )

    # Write totals
    first_data_col = ul_col + 2
    last_data_col = first_data_col + len(classes) - 1
    first_data_col_letter = get_column_letter(first_data_col)
    last_data_col_letter = get_column_letter(last_data_col)
    first_data_row = ul_row + 3
    last_data_row = first_data_row + len(classes) - 1
    for n in range(len(classes)):
        # Write row total
        row_total_cell = sheet.cell(row=first_data_row + n, column=last_data_col + 1)
        row_total_cell.value = (
            f"=sum({first_data_col_letter}{first_data_row + n}"
            + f":{last_data_col_letter}{first_data_row + n})"
        )
        row_total_cell.font = Font(italic=True)
        row_total_cell.alignment = Alignment(horizontal="center")
        row_total_cell.number_format = "#,##0.00"
        # Write colum total
        col_total_cell = sheet.cell(row=last_data_row + 1, column=first_data_col + n)
        col_total_cell.value = (
            f"=sum({get_column_letter(first_data_col + n)}{first_data_row}"
            + f":{get_column_letter(first_data_col + n)}{last_data_row})"
        )
        col_total_cell.font = Font(italic=True)
        col_total_cell.alignment = Alignment(horizontal="center")
        col_total_cell.number_format = "#,##0.00"

    # Add bottom right corner total
    br_cell = sheet.cell(row=last_data_row + 1, column=last_data_col + 1)
    br_cell.value = (
        f"=sum({first_data_col_letter}{last_data_row + 1}"
        + f":{last_data_col_letter}{last_data_row + 1})"
    )
    br_cell.font = Font(italic=True, bold=True)
    br_cell.alignment = Alignment(horizontal="center")
    br_cell.number_format = "#,##0.00"

    # Merge main header
    sheet.merge_cells(
        start_row=ul_row, start_column=1, end_row=ul_row, end_column=last_data_col + 1
    )
    # Merge table row/column headers
    sheet.merge_cells(
        start_row=ul_row + 1,
        start_column=first_data_col,
        end_row=ul_row + 1,
        end_column=last_data_col,
    )
    sheet.merge_cells(
        start_row=first_data_row,
        start_column=ul_col,
        end_row=last_data_row,
        end_column=ul_col,
    )


def _write_land_cover_sheet(
    sheet,
    st: models.SummaryTableLD,
    lc_trans_matrix: land_cover.LCTransitionDefinitionDeg,
    period,
):
    classes = [c.get_name() for c in lc_trans_matrix.legend.key]
    if len(classes) > 7:
        sheet.insert_rows(19 + 1, len(classes) - 7)
        n_new_rows_per_table = len(classes) - 7
    else:
        n_new_rows_per_table = 0

    xl.write_col_to_sheet(sheet, _get_summary_array(st.lc_summary), 6, 6)

    # Write land cover change by class table
    first_data_row = 14
    last_data_row = first_data_row + len(classes) - 1
    xl.write_col_to_sheet(
        sheet, classes, 2, first_data_row, header=True, border=True, wrap=True
    )
    for row in sheet.iter_rows(
        min_row=first_data_row, max_row=last_data_row, min_col=3, max_col=6
    ):
        for cell in row:
            if cell.column == 3:
                # Initial area column
                cell.value = (
                    f"={get_column_letter(cell.column + len(classes))}"
                    + f"{cell.offset(row=12 + n_new_rows_per_table).row}"
                )
                cell.number_format = "#,##0.00"
            elif cell.column == 4:
                # Final area column
                col_letter = cell.offset(
                    column=cell.row - first_data_row - 1
                ).column_letter
                cell.value = f"={col_letter}{33 + n_new_rows_per_table * 2}"
                cell.number_format = "#,##0.00"
            elif cell.column == 5:
                # Change in area column
                cell.value = (
                    f"={cell.offset(column=-1).column_letter}{cell.row} - "
                    + f"{cell.offset(column=-2).column_letter}{cell.row}"
                )
                cell.number_format = "#,##0.00"
            elif cell.column == 6:
                # Percent change in area column
                cell.value = (
                    f"={cell.offset(column=-1).column_letter}{cell.row} / "
                    + f"{cell.offset(column=-3).column_letter}{cell.row}"
                )
                cell.number_format = numbers.FORMAT_PERCENTAGE
            cell.alignment = Alignment(horizontal="center")
            cell.border = xl.thin_border

    # Add total rows
    for row in sheet.iter_rows(
        min_row=last_data_row + 1, max_row=last_data_row + 1, min_col=3, max_col=4
    ):
        for cell in row:
            cell.value = (
                f"=sum({cell.column_letter}{first_data_row}"
                + f":{cell.column_letter}{last_data_row})"
            )
            cell.font = Font(italic=True)
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = "#,##0.00"

    lc_trans_zonal_areas = [
        x
        for x, p in zip(st.lc_trans_zonal_areas, st.lc_trans_zonal_areas_periods)
        if p == period
    ][0]

    write_crosstab_by_lc(
        sheet,
        _get_lc_trans_table(lc_trans_zonal_areas, lc_trans_matrix),
        classes,
        23 + n_new_rows_per_table,
        1,
    )

    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet)


def _write_population_sheet(sheet, st: models.SummaryTableLD):
    xl.write_col_to_sheet(
        sheet, _get_summary_array(st.sdg_zonal_population_total), 6, 6
    )
    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet)


def _get_prod_table(lc_trans_prod_bizonal, prod_code, lc_trans_matrix):
    lc_codes = sorted([c.code for c in lc_trans_matrix.legend.key])
    out = np.zeros((len(lc_codes), len(lc_codes)))

    transition_key = lc_trans_matrix.get_transition_initial_final_key()

    for i, i_code in enumerate(lc_codes):
        for f, f_code in enumerate(lc_codes):
            transition_code = transition_key[i_code][f_code]
            out[i, f] = lc_trans_prod_bizonal.get((transition_code, prod_code), 0.0)

    return out


def _get_totals_by_lc_class_as_array(
    annual_totals,
    lc_trans_matrix,
    excluded_codes=[],  # to exclude water when used on SOC table
):
    lc_codes = [
        lc_trans_matrix.legend.class_index(c)
        for c in sorted(lc_trans_matrix.legend.key, key=lambda i: i.code)
        if c.code not in excluded_codes
    ]

    return np.array([annual_totals.get(lc_code, 0.0) for lc_code in lc_codes])


def _write_soc_stock_change_table(
    sheet,
    ul_row,
    ul_col,
    soc_bl_totals,
    soc_final_totals,
    lc_trans_matrix,
    excluded_codes=[],  # to exclude water
):
    first_row = ul_row + 3
    first_col = ul_col + 2

    lc_codes = sorted(
        [c.code for c in lc_trans_matrix.legend.key if c.code not in excluded_codes]
    )

    transition_key = lc_trans_matrix.get_transition_initial_final_key()

    for i, i_code in enumerate(lc_codes):
        for f, f_code in enumerate(lc_codes):
            cell = sheet.cell(row=i + first_row, column=f + first_col)
            transition_code = transition_key[i_code][f_code]
            bl_soc = soc_bl_totals.get(transition_code, 0.0)
            final_soc = soc_final_totals.get(transition_code, 0.0)
            cell.border = xl.thin_border
            cell.alignment = Alignment(horizontal="center")
            try:
                cell.value = (final_soc - bl_soc) / bl_soc
                cell.number_format = "#,##0.00"
            except ZeroDivisionError:
                cell.value = ""


def _get_lc_trans_table(lc_trans_totals, lc_trans_matrix, excluded_codes=[]):
    lc_codes = sorted(
        [c.code for c in lc_trans_matrix.legend.key if c.code not in excluded_codes]
    )
    out = np.zeros((len(lc_codes), len(lc_codes)))

    transition_key = lc_trans_matrix.get_transition_initial_final_key()

    for i, i_code in enumerate(lc_codes):
        for f, f_code in enumerate(lc_codes):
            transition_code = transition_key[i_code][f_code]
            out[i, f] = lc_trans_totals.get(transition_code, 0.0)

    return out


def _get_soc_total(soc_table, transition):
    ind = np.where(soc_table[0] == transition)[0]

    if ind.size == 0:
        return 0
    else:
        return float(soc_table[1][ind])
