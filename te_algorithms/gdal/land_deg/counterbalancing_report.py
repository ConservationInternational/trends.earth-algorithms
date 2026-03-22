"""Reporting for LDN counterbalancing assessment.

Generates:
  - Excel workbook with counterbalancing summary and land type achievement
    sheets.
  - JSON report following the ``save_reporting_json`` pattern.
"""

import datetime as dt
import json
import logging
from pathlib import Path
from typing import TYPE_CHECKING, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.cell import get_column_letter

from ... import __release_date__, __version__
from .. import xl
from . import models

if TYPE_CHECKING:
    from te_schemas.aoi import AOI

logger = logging.getLogger(__name__)

_header_font = Font(bold=True)
_header_alignment = Alignment(horizontal="center", wrap_text=True)
_center_alignment = Alignment(horizontal="center")


def _add_header_cell(sheet, row, col, value):
    cell = sheet.cell(row=row, column=col)
    cell.value = value
    cell.alignment = _header_alignment
    cell.border = xl.thin_border
    cell.font = _header_font


def _add_data_cell(sheet, row, col, value, fmt=None):
    cell = sheet.cell(row=row, column=col)
    cell.value = value
    cell.alignment = _center_alignment
    cell.border = xl.thin_border
    if fmt:
        cell.number_format = fmt


def save_counterbalancing_excel(
    output_path: Path,
    land_type_results: List[models.CounterbalancingLandTypeResult],
    summary_table: models.SummaryTableCounterbalancing,
):
    """Save counterbalancing results into an Excel workbook."""
    wb = Workbook()

    _write_summary_sheet(wb.active, land_type_results)

    try:
        wb.save(str(output_path))
        logger.info("Counterbalancing table saved to %s", output_path)
    except OSError:
        logger.error(
            "Error saving counterbalancing table - check that %s is accessible "
            "and not already open.",
            output_path,
        )


def _write_summary_sheet(
    sheet, land_type_results: List[models.CounterbalancingLandTypeResult]
):
    """Overview sheet: gains, losses, delta per land type."""
    sheet.title = "Counterbalancing Summary"
    xl.maybe_add_image_to_sheet("trends_earth_logo_bl_300width.png", sheet)

    # Title
    sheet.cell(row=1, column=1, value="LDN Counterbalancing Summary").font = Font(
        bold=True, size=14
    )

    # Headers at row 4
    headers = [
        "Land Type",
        "Gains (sq km)",
        "Losses (sq km)",
        "\u0394 LDN (sq km)",
        "LDN Achieved",
    ]
    for col, h in enumerate(headers, 1):
        _add_header_cell(sheet, 4, col, h)

    # Data rows starting at row 5
    row = 5
    total_gains = 0.0
    total_losses = 0.0

    for lt in sorted(land_type_results, key=lambda r: r.land_type_code):
        total_gains += lt.gains_area_sqkm
        total_losses += lt.losses_area_sqkm

        sheet.cell(row=row, column=1, value=lt.land_type_name).border = xl.thin_border
        _add_data_cell(sheet, row, 2, lt.gains_area_sqkm, "#,##0.00")
        _add_data_cell(sheet, row, 3, lt.losses_area_sqkm, "#,##0.00")
        _add_data_cell(sheet, row, 4, lt.delta_ldn, "#,##0.00")
        cell = sheet.cell(row=row, column=5)
        cell.value = "Yes" if lt.ldn_achieved else "No"
        cell.alignment = _center_alignment
        cell.border = xl.thin_border
        cell.font = Font(bold=True, color="006500" if lt.ldn_achieved else "9b2779")
        row += 1

    # Totals row
    total_delta = total_gains - total_losses
    sheet.cell(row=row, column=1, value="Total").font = Font(bold=True, italic=True)
    sheet.cell(row=row, column=1).border = xl.thin_border
    _add_data_cell(sheet, row, 2, total_gains, "#,##0.00")
    sheet.cell(row=row, column=2).font = Font(italic=True, bold=True)
    _add_data_cell(sheet, row, 3, total_losses, "#,##0.00")
    sheet.cell(row=row, column=3).font = Font(italic=True, bold=True)
    _add_data_cell(sheet, row, 4, total_delta, "#,##0.00")
    sheet.cell(row=row, column=4).font = Font(italic=True, bold=True)
    cell = sheet.cell(row=row, column=5)
    all_achieved = total_delta >= 0
    cell.value = "Yes" if all_achieved else "No"
    cell.alignment = _center_alignment
    cell.border = xl.thin_border
    cell.font = Font(
        bold=True, italic=True, color="006500" if all_achieved else "9b2779"
    )

    # Auto-size columns
    for col in range(1, 6):
        sheet.column_dimensions[get_column_letter(col)].width = 18


def save_counterbalancing_json(
    output_path: Path,
    land_type_results: List[models.CounterbalancingLandTypeResult],
    task_name: str,
    aoi: "AOI",
    land_type_labels: dict = None,
    land_type_layer_paths: list = None,
):
    """Save counterbalancing results as JSON and return the report dict."""
    try:
        version = __version__
    except Exception:
        version = "unknown"

    try:
        release_date = __release_date__
    except Exception:
        release_date = "unknown"

    report = {
        "metadata": {
            "algorithm_version": version,
            "release_date": release_date,
            "report_date": dt.datetime.now(dt.timezone.utc).isoformat(),
            "task_name": task_name,
        },
        "land_type_results": [
            {
                "land_type_code": lt.land_type_code,
                "land_type_name": lt.land_type_name,
                "gains_area_sqkm": lt.gains_area_sqkm,
                "losses_area_sqkm": lt.losses_area_sqkm,
                "delta_ldn": lt.delta_ldn,
                "ldn_achieved": lt.ldn_achieved,
            }
            for lt in land_type_results
        ],
    }

    # Add spatial unit key so users can decode which original layer values
    # produced each land type code in the spatial units raster.
    if land_type_labels is not None:
        layer_names = [Path(p).stem for p in (land_type_layer_paths or [])]
        report["spatial_unit_key"] = {
            "description": (
                "Maps each spatial unit code in the land type raster to the "
                "combination of values from the input layers that define it. "
                "The 'values' list is ordered to match 'layer_names'."
            ),
            "layer_names": layer_names,
            "units": {
                str(code): {
                    "label": label,
                    "values": label.split("_"),
                }
                for code, label in sorted(land_type_labels.items())
            },
        }

    try:
        with open(str(output_path), "w") as f:
            json.dump(report, f, indent=2)
        logger.info("Counterbalancing JSON report saved to %s", output_path)
    except OSError:
        logger.error(
            "Error saving counterbalancing JSON - check that %s is accessible",
            output_path,
        )

    return report
